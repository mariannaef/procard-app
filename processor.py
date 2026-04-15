import pdfplumber
import pandas as pd
import zipfile
import re
import argparse
from io import BytesIO
from pathlib import Path
from difflib import SequenceMatcher
from datetime import datetime, date
from openpyxl.styles import Font

STATE_CODES = {
    'AL', 'AK', 'AZ', 'AR', 'CA', 'CO', 'CT', 'DE', 'FL', 'GA', 'HI', 'ID', 'IL', 'IN', 'IA',
    'KS', 'KY', 'LA', 'ME', 'MD', 'MA', 'MI', 'MN', 'MS', 'MO', 'MT', 'NE', 'NV', 'NH', 'NJ',
    'NM', 'NY', 'NC', 'ND', 'OH', 'OK', 'OR', 'PA', 'RI', 'SC', 'SD', 'TN', 'TX', 'UT', 'VT',
    'VA', 'WA', 'WV', 'WI', 'WY'
}

GENERIC_MERCHANT_TOKENS = {'HTTP', 'HTTPS', 'WWW', 'COM', 'PROD', 'HELP', 'GOSQ'}
FILE_FEED_COLUMN_WIDTHS = [31.14, 1.57, 12.43, 1.29, 28.14, 6.29, 5.29]
WORKBOOK_FONT_NAME = 'Aptos Narrow'
PROCARD_TOTAL_TASK_NAME = 'PROCARD TOTAL'
PROCARD_TOTAL_LABEL = 'PROCARD'


def apply_workbook_font(workbook, font_name=WORKBOOK_FONT_NAME, font_size=11):
    """Apply default workbook font (including Normal style) across all sheets."""
    font = Font(name=font_name, size=font_size)

    # Set workbook default "Normal" style so Excel width rendering is consistent.
    try:
        for named_style in getattr(workbook, "_named_styles", []):
            if getattr(named_style, "name", "") == "Normal":
                named_style.font = font
                break
    except Exception:
        pass

    for ws in workbook.worksheets:
        if ws.max_row <= 0 or ws.max_column <= 0:
            continue
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = font

def clean_val(val):
    """Cleans tabs and whitespace from CSV data."""
    return str(val).replace('\t', '').strip() if pd.notna(val) else ""


def parse_amount(val, default=0.0):
    """Safely parse amount-like values from CSV/PDF fields."""
    s = clean_val(val).replace('$', '').replace(',', '')
    if s == "":
        return default
    try:
        return float(s)
    except ValueError:
        return default


def procard_total_row_mask(df):
    """Identify synthetic PROCARD total rows by task name."""
    if df is None or len(df) == 0:
        return pd.Series(dtype=bool)
    if 'Current Task Name' not in df.columns:
        return pd.Series([False] * len(df), index=df.index)
    return df['Current Task Name'].apply(lambda v: clean_val(v).upper() == PROCARD_TOTAL_TASK_NAME)


def append_procard_total_row(teams_workflow, total_amount=None):
    """Append a bottom summary row labeled PROCARD with Statement Credit=True.

    Credits remain negative and debits positive in the net total amount.
    """
    if teams_workflow is None:
        return teams_workflow

    result = teams_workflow.copy()
    existing_mask = procard_total_row_mask(result)
    if len(existing_mask) == len(result) and existing_mask.any():
        result = result.loc[~existing_mask].copy()

    if total_amount is None:
        total_amount = float(round(result.get('Amount: $', pd.Series(dtype=float)).apply(parse_amount).sum(), 2))
    else:
        total_amount = float(round(parse_amount(total_amount), 2))

    summary_row = {col: "" for col in result.columns}
    if 'Current Task Name' in summary_row:
        summary_row['Current Task Name'] = PROCARD_TOTAL_TASK_NAME
    if 'Sport' in summary_row:
        summary_row['Sport'] = PROCARD_TOTAL_LABEL
    if 'Merchant:' in summary_row:
        summary_row['Merchant:'] = PROCARD_TOTAL_LABEL
    if 'Date of Transaction:' in summary_row:
        summary_row['Date of Transaction:'] = ""
    if 'Amount: $' in summary_row:
        summary_row['Amount: $'] = total_amount
    if 'Card Holder:' in summary_row:
        summary_row['Card Holder:'] = PROCARD_TOTAL_LABEL
    if 'Fund Code:' in summary_row:
        summary_row['Fund Code:'] = ""
    if 'Organization:' in summary_row:
        summary_row['Organization:'] = ""
    if 'Account:' in summary_row:
        summary_row['Account:'] = ""
    if 'Program:' in summary_row:
        summary_row['Program:'] = ""
    if 'AD Code:' in summary_row:
        summary_row['AD Code:'] = ""
    if 'Amount:' in summary_row:
        summary_row['Amount:'] = ""
    if 'Statement Credit' in summary_row:
        summary_row['Statement Credit'] = True

    return pd.concat([result, pd.DataFrame([summary_row], columns=result.columns)], ignore_index=True)


def normalize_mmdd(val):
    """Normalize date-like text to MM/DD for reliable matching."""
    s = clean_val(val)
    if not s:
        return ""

    m = re.search(r'^(\d{1,2})/(\d{1,2})(?:/\d{2,4})?$', s)
    if not m:
        # Try to locate first M/D pattern inside noisy text.
        m = re.search(r'\b(\d{1,2})/(\d{1,2})\b', s)
    if not m:
        return ""

    month = int(m.group(1))
    day = int(m.group(2))
    if month < 1 or month > 12 or day < 1 or day > 31:
        return ""
    return f"{month:02d}/{day:02d}"


def extract_statement_cycle_range(first_page_text):
    """Extract statement cycle date range from first page text (e.g., Feb 18 - Mar 17, 2026)."""
    text = clean_val(first_page_text)
    if not text:
        return None, None

    match = re.search(r'\b([A-Za-z]{3})\s+(\d{1,2})\s*-\s*([A-Za-z]{3})\s+(\d{1,2}),\s*(\d{4})\b', text)
    if not match:
        return None, None

    start_mon, start_day, end_mon, end_day, end_year = match.groups()
    try:
        end_date = datetime.strptime(f"{end_mon} {end_day} {end_year}", "%b %d %Y").date()
        start_date = datetime.strptime(f"{start_mon} {start_day} {end_year}", "%b %d %Y").date()
        if start_date > end_date:
            start_date = datetime.strptime(f"{start_mon} {start_day} {int(end_year) - 1}", "%b %d %Y").date()
        return start_date, end_date
    except ValueError:
        return None, None


def parse_workflow_date(value, cycle_start=None, cycle_end=None):
    """Parse workflow date to a real date object, using cycle years when year is missing."""
    s = clean_val(value)
    if not s:
        return None

    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass

    md = re.search(r'\b(\d{1,2})/(\d{1,2})\b', s)
    if not md:
        return None

    month = int(md.group(1))
    day = int(md.group(2))
    candidate_years = []
    if cycle_start is not None:
        candidate_years.append(cycle_start.year)
    if cycle_end is not None and cycle_end.year not in candidate_years:
        candidate_years.append(cycle_end.year)
    if not candidate_years:
        candidate_years.append(datetime.now().year)

    candidates = []
    for y in candidate_years:
        try:
            candidates.append(date(y, month, day))
        except ValueError:
            continue
    if not candidates:
        return None

    if cycle_start is not None and cycle_end is not None:
        for c in candidates:
            if cycle_start <= c <= cycle_end:
                return c
        # closest to cycle end when outside range
        return min(candidates, key=lambda d: abs((d - cycle_end).days))

    return candidates[0]


def format_statement_cycle(cycle_start, cycle_end):
    """Format statement cycle bounds as MM/DD/YYYY - MM/DD/YYYY."""
    if not cycle_start or not cycle_end:
        return ""
    return f"{cycle_start.strftime('%m/%d/%Y')} - {cycle_end.strftime('%m/%d/%Y')}"


def classify_date_outside_cycle(parsed_date, cycle_start, cycle_end):
    """Return the same warning text used for workflow rows outside the statement cycle."""
    if parsed_date is None or not (cycle_start and cycle_end):
        return ""

    cycle_range = format_statement_cycle(cycle_start, cycle_end)
    if parsed_date < cycle_start:
        return f"Likely on previous statement (outside {cycle_range})"
    if parsed_date > cycle_end:
        return f"Likely on next statement (outside {cycle_range})"
    return ""


def expand_split_transactions(workflow):
    """Expand workflow rows with multiple FOAPAL distribution lines into separate rows.

    Column behaviour after expansion:
      'Amount: $'  — main row keeps the overall transaction total; split rows are blank.
      'Amount:'    — every row (main + splits) holds its own distribution amount.
                     Main row uses 'Amount:' as-is; split rows pull from 'Amount:.1', etc.
      FOAPAL codes — each row gets its own codes from the numbered columns (.1, .2, …).

    Split column naming pattern in the CSV:
      Line 1 (main): 'Fund Code:',   'Organization:',   ..., 'Amount:'
      Line 2:        'Fund Code:.1', 'Organization:.1', ..., 'Amount:.1'
      Line 3:        'Fund Code:.2', 'Organization:.2', ..., 'Amount:.2'
      etc.
    """
    foapal_cols = ['Fund Code:', 'Organization:', 'Account:', 'Program:', 'AD Code:', 'Amount:']
    expanded_rows = []

    def clean_foapal_val(val):
        """Like clean_val but also converts whole-number floats (e.g. 209030.0) to int strings."""
        s = clean_val(val)
        if s:
            try:
                f = float(s)
                if f == int(f):
                    return str(int(f))
            except (ValueError, TypeError):
                pass
        return s

    for _, row in workflow.iterrows():
        num_lines_str = row.get('How many transaction lines would you like to have listed?', '')
        num_lines = 1
        if pd.notna(num_lines_str) and str(num_lines_str).strip():
            try:
                num_lines = int(float(str(num_lines_str).strip()))
            except (ValueError, TypeError):
                num_lines = 1

        # Always add the main row (Amount: $ = total, Amount: = first distribution).
        # For non-split rows, Amount: should mirror Amount: $ when distribution is missing/zero.
        main_row = row.copy()
        if num_lines <= 1:
            main_distribution = parse_amount(main_row.get('Amount:', 0.0))
            main_total = parse_amount(main_row.get('Amount: $', 0.0))
            if main_distribution == 0.0 and main_total != 0.0:
                main_row['Amount:'] = main_row.get('Amount: $')
        expanded_rows.append(main_row)

        # Add one row per additional distribution line (split_idx 1, 2, 3, 4)
        for split_idx in range(1, num_lines):
            split_row = row.copy()
            split_row['Amount: $'] = None   # total is blank on split rows
            for col in foapal_cols:
                split_col = f'{col}.{split_idx}'   # e.g. 'Fund Code:.1', 'Amount:.1'
                val = row.get(split_col) if split_col in workflow.columns else None
                # Parse amount strings; strip tabs/whitespace and normalise floats for FOAPAL values
                split_row[col] = parse_amount(val) if col == 'Amount:' else clean_foapal_val(val)
            expanded_rows.append(split_row)

    return pd.DataFrame(expanded_rows).reset_index(drop=True)


def write_output_workbook(output_file, file_feed, teams_workflow, bank_review, workflow_review, workflow_only_rows):
    """Write the final workbook sheets in the established order and format."""
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        file_feed.to_excel(writer, sheet_name='FILE FEED', index=False, header=False)
        teams_workflow.to_excel(writer, sheet_name='TEAMS WORKFLOW', index=False)
        bank_review.to_excel(writer, sheet_name='PDF Parsed Transactions', index=False)
        workflow_review.to_excel(writer, sheet_name='Original Workflow', index=False)
        workflow_only_rows.to_excel(writer, sheet_name='Workflow Not In PDF', index=False)

        file_feed_ws = writer.sheets.get('FILE FEED')
        if file_feed_ws is not None:
            for col_letter, width in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G'], FILE_FEED_COLUMN_WIDTHS):
                file_feed_ws.column_dimensions[col_letter].width = width
        apply_workbook_font(writer.book)


def build_file_feed_from_teams_workflow(teams_workflow, batch_date):
    """Build FILE FEED dataframe from TEAMS WORKFLOW dataframe."""
    if teams_workflow is None or teams_workflow.empty:
        return pd.DataFrame(columns=['Col1', 'Col2', 'Col3', 'Col4', 'Col5', 'Col6', 'Col7'])

    summary_mask = procard_total_row_mask(teams_workflow)
    summary_rows = teams_workflow.loc[summary_mask].copy() if len(summary_mask) == len(teams_workflow) else pd.DataFrame(columns=teams_workflow.columns)
    teams_workflow = teams_workflow.loc[~summary_mask].copy() if len(summary_mask) == len(teams_workflow) else teams_workflow.copy()

    def format_segment(val):
        s = str(val).split('.')[0].strip() if pd.notna(val) else ""
        return s.ljust(6)[:6]

    def build_col1_from_row(row):
        return (
            format_segment(row.get('Fund Code:', '')) +
            format_segment(row.get('Organization:', '')) +
            format_segment(row.get('Account:', '')) +
            format_segment(row.get('Program:', '')) +
            format_segment(row.get('AD Code:', ''))
        )

    def row_has_real_foapal(row):
        for col in ['Fund Code:', 'Organization:', 'Account:', 'Program:', 'AD Code:']:
            token = clean_val(row.get(col, '')).upper()
            if token and token not in {PROCARD_TOTAL_LABEL, PROCARD_TOTAL_TASK_NAME}:
                return True
        return False

    def merchant_with_leading_tab(val):
        merchant = clean_val(val)
        return f"\t{merchant}" if merchant else ""

    if teams_workflow.empty:
        file_feed = pd.DataFrame(columns=['Col1', 'Col2', 'Col3', 'Col4', 'Col5', 'Col6', 'Col7'])
    else:
        def effective_distribution_amount(row):
            """Use split amount when present; otherwise retain original transaction amount."""
            split_amount_raw = row.get('Amount:', '')
            split_amount = parse_amount(split_amount_raw)
            transaction_amount = parse_amount(row.get('Amount: $', 0.0))

            # In source data, blank distribution amounts can be coerced to 0.0.
            # When that happens, keep the transaction amount so FILE FEED Col3 is not zeroed out.
            if clean_val(split_amount_raw) == "" or (split_amount == 0.0 and transaction_amount != 0.0):
                return transaction_amount
            return split_amount

        file_feed = pd.DataFrame()
        # 30-character account string
        file_feed['Col1'] = (
            teams_workflow['Fund Code:'].apply(format_segment) +
            teams_workflow['Organization:'].apply(format_segment) +
            teams_workflow['Account:'].apply(format_segment) +
            teams_workflow['Program:'].apply(format_segment) +
            teams_workflow['AD Code:'].apply(format_segment)
        )
        file_feed['Col2'] = ""
        effective_amounts = teams_workflow.apply(effective_distribution_amount, axis=1)
        # 12-digit cents
        file_feed['Col3'] = effective_amounts.apply(
            lambda x: f"{int(round(abs(x) * 100)):012d}"
        )
        file_feed['Col4'] = teams_workflow.assign(_effective_amount=effective_amounts).apply(
            lambda row: '-' if row['Statement Credit'] and row['_effective_amount'] < 0 else (
                '-' if row['_effective_amount'] < 0 else '+'
            ),
            axis=1
        )
        file_feed['Col5'] = teams_workflow['Merchant:'].apply(merchant_with_leading_tab)
        file_feed['Col6'] = batch_date
        file_feed['Col7'] = '*****'

    # Add PROCARD total row to FILE FEED bottom, labeled as credit.
    if not summary_rows.empty:
        summary_row = summary_rows.iloc[-1]
        summary_amount = parse_amount(summary_row.get('Amount: $', 0.0))

        summary_col1 = ""
        if row_has_real_foapal(summary_row):
            summary_col1 = build_col1_from_row(summary_row)

        summary_col5 = clean_val(summary_row.get('Merchant:', '')) or 'ProCard'
        if summary_col5.upper() == PROCARD_TOTAL_LABEL:
            summary_col5 = 'ProCard'

        file_feed = pd.concat([
            file_feed,
            pd.DataFrame([{
                'Col1': summary_col1,
                'Col2': '',
                'Col3': f"{int(round(abs(summary_amount) * 100)):012d}",
                'Col4': '-',
                'Col5': merchant_with_leading_tab(summary_col5),
                'Col6': batch_date,
                'Col7': '*****',
            }])
        ], ignore_index=True)

    return file_feed


def is_blank(val):
    """True when a value is empty/NaN after cleanup."""
    s = clean_val(val)
    return s == "" or s.lower() == "nan"


def prompt_foapal_values():
    """Prompt once for all FOAPAL fields and return validated values."""
    while True:
        raw = clean_val(input("Enter FOAPAL as Fund,Organization,Account,Program,AD Code: "))
        if not raw:
            print("FOAPAL is required.")
            continue

        parts = [clean_val(part) for part in raw.split(',')]
        if len(parts) != 5 or any(not part for part in parts):
            print("Please enter exactly 5 comma-separated values: Fund,Organization,Account,Program,AD Code")
            continue

        return {
            'Fund Code:': parts[0],
            'Organization:': parts[1],
            'Account:': parts[2],
            'Program:': parts[3],
            'AD Code:': parts[4],
        }


def fill_missing_foapal(teams_workflow):
    """Prompt for FOAPAL when Fund Code is missing, once per purchase key."""
    required_cols = ['Fund Code:', 'Organization:', 'Account:', 'Program:', 'AD Code:']

    # Ensure prompted string values can be assigned even if source columns were numeric
    teams_workflow[required_cols] = teams_workflow[required_cols].astype(object)

    missing_mask = teams_workflow['Fund Code:'].apply(is_blank)

    if not missing_mask.any():
        return teams_workflow

    print("\nMissing accounting codes detected. Please enter FOAPAL values for each purchase.")
    purchase_cache = {}

    for idx in teams_workflow[missing_mask].index:
        purchase_key = (
            clean_val(teams_workflow.at[idx, 'Card Holder:']),
            clean_val(teams_workflow.at[idx, 'Merchant:']),
            clean_val(teams_workflow.at[idx, 'Date of Transaction:']),
            f"{parse_amount(teams_workflow.at[idx, 'Amount: $']):.2f}"
        )

        if purchase_key not in purchase_cache:
            card_holder, merchant, tx_date, tx_amount = purchase_key
            sport = clean_val(teams_workflow.at[idx, 'Sport'])
            print(f"\nPurchase: {tx_date} | {merchant} | ${tx_amount} | {card_holder} | {sport}")
            purchase_cache[purchase_key] = prompt_foapal_values()

        for col in required_cols:
            if is_blank(teams_workflow.at[idx, col]):
                teams_workflow.at[idx, col] = purchase_cache[purchase_key][col]

    return teams_workflow


def first_matching_column(df, column_name):
    """Return the first physical column for a duplicated CSV header name."""
    matches = [index for index, name in enumerate(df.columns) if name == column_name]
    if not matches:
        raise KeyError(f"Missing required column: {column_name}")
    return df.iloc[:, matches[0]]


def normalize_match_text(value):
    """Normalize merchant/transaction text for robust reconciliation matching."""
    text = clean_val(value).upper()
    text = re.sub(r'https?://\S+|www\.\S+', ' ', text)
    text = re.sub(r'([A-Z])([0-9])', r'\1 \2', text)
    text = re.sub(r'([0-9])([A-Z])', r'\1 \2', text)
    text = re.sub(r'[^A-Z0-9]+', ' ', text)

    tokens = []
    for token in text.split():
        if any(ch.isdigit() for ch in token):
            continue
        if token in STATE_CODES or token in GENERIC_MERCHANT_TOKENS:
            continue
        tokens.append(token)

    return ' '.join(tokens)


def merchant_match_score(left_text, right_text):
    """Score similarity between parsed transaction text and workflow merchant text."""
    left_norm = normalize_match_text(left_text)
    right_norm = normalize_match_text(right_text)
    if not left_norm or not right_norm:
        return 0.0

    if left_norm == right_norm:
        return 1.0

    left_tokens = set(left_norm.split())
    right_tokens = set(right_norm.split())
    overlap = len(left_tokens & right_tokens) / max(1, min(len(left_tokens), len(right_tokens)))
    sequence = SequenceMatcher(None, left_norm, right_norm).ratio()
    contains = 1.0 if left_norm in right_norm or right_norm in left_norm else 0.0
    return max(overlap, sequence, contains)


def relaxed_merchant_match_score(tx_date, tx_text, workflow_merchant, workflow_date):
    """Amount-match fallback score with optional day-proximity bonus."""
    base = merchant_match_score(tx_text, workflow_merchant)

    tx_day = None
    try:
        tx_day = int(tx_date.split('/')[1]) if tx_date else None
    except Exception:
        tx_day = None

    row_date = normalize_mmdd(workflow_date)
    day_bonus = 0.0
    if tx_day is not None and row_date:
        try:
            row_day = int(row_date.split('/')[1])
            if abs(row_day - tx_day) <= 2:
                day_bonus = 0.03
        except Exception:
            pass

    return min(1.0, base + day_bonus)


def find_best_workflow_match(workflow, tx_date, tx_amount, tx_text, matched_workflow_rows, threshold=0.72):
    """Return the best unmatched workflow row for a parsed transaction."""
    candidates = workflow[
        (workflow['_Date_MMDD'] == tx_date) &
        (workflow['Amount: $'].round(2) == tx_amount)
    ]

    if matched_workflow_rows:
        candidates = candidates.loc[~candidates.index.isin(matched_workflow_rows)]

    if candidates.empty:
        return pd.DataFrame(), 0.0

    scored_candidates = candidates.copy()
    scored_candidates['_Match_Score'] = scored_candidates['Merchant:'].apply(
        lambda merchant: merchant_match_score(tx_text, merchant)
    )
    scored_candidates = scored_candidates.loc[scored_candidates['_Match_Score'] >= threshold]
    if scored_candidates.empty:
        return pd.DataFrame(), 0.0

    best_index = scored_candidates['_Match_Score'].idxmax()
    best_row = scored_candidates.loc[[best_index]].copy()
    return best_row, float(best_row.iloc[0]['_Match_Score'])


def find_best_workflow_match_relaxed(workflow, tx_date, tx_amount, tx_text, matched_workflow_rows, threshold=0.93):
    """Fallback match using amount + strong text similarity (date can drift)."""
    candidates = workflow[(workflow['Amount: $'].round(2) == tx_amount)].copy()
    if matched_workflow_rows:
        candidates = candidates.loc[~candidates.index.isin(matched_workflow_rows)]
    if candidates.empty:
        return pd.DataFrame(), 0.0

    candidates['_Relaxed_Match_Score'] = candidates.apply(
        lambda row: relaxed_merchant_match_score(
            tx_date,
            tx_text,
            row.get('Merchant:', ''),
            row.get('Date of Transaction:', ''),
        ),
        axis=1,
    )
    candidates = candidates.loc[candidates['_Relaxed_Match_Score'] >= threshold]
    if candidates.empty:
        return pd.DataFrame(), 0.0

    best_index = candidates['_Relaxed_Match_Score'].idxmax()
    best_row = candidates.loc[[best_index]].copy()
    return best_row, float(best_row.iloc[0]['_Relaxed_Match_Score'])


def get_workflow_match_candidates(
    workflow,
    tx_date,
    tx_amount,
    tx_text,
    matched_workflow_rows,
    threshold=0.55,
    relaxed=False,
    max_candidates=8,
):
    """Return scored candidate workflow rows for manual review."""
    if relaxed:
        candidates = workflow[(workflow['Amount: $'].round(2) == tx_amount)].copy()
    else:
        candidates = workflow[
            (workflow['_Date_MMDD'] == tx_date) &
            (workflow['Amount: $'].round(2) == tx_amount)
        ].copy()

    if matched_workflow_rows:
        candidates = candidates.loc[~candidates.index.isin(matched_workflow_rows)]

    if candidates.empty:
        return candidates

    if relaxed:
        candidates['_Match_Score'] = candidates.apply(
            lambda row: relaxed_merchant_match_score(
                tx_date,
                tx_text,
                row.get('Merchant:', ''),
                row.get('Date of Transaction:', ''),
            ),
            axis=1,
        )
    else:
        candidates['_Match_Score'] = candidates['Merchant:'].apply(
            lambda merchant: merchant_match_score(tx_text, merchant)
        )

    candidates = candidates.loc[candidates['_Match_Score'] >= threshold]
    if candidates.empty:
        return candidates

    return candidates.sort_values('_Match_Score', ascending=False).head(max_candidates)


def prompt_user_to_approve_match(tx_row, candidates):
    """Prompt user to approve one candidate match or skip."""
    if candidates.empty:
        return None

    tx_date = normalize_mmdd(tx_row.get('Date', ''))
    tx_amount = round(parse_amount(tx_row.get('Amount', 0.0)), 2)
    tx_transaction = clean_val(tx_row.get('Transaction', ''))
    tx_holder = clean_val(tx_row.get('Card Holder', ''))

    print("\nPossible duplicate/alternate matches found:")
    print(f"  Bank Tx: {tx_date} | {tx_transaction} | ${tx_amount:.2f} | {tx_holder}")
    for option_num, (idx, row) in enumerate(candidates.iterrows(), start=1):
        w_date = normalize_mmdd(row.get('Date of Transaction:', ''))
        w_merchant = clean_val(row.get('Merchant:', ''))
        w_holder = clean_val(row.get('Card Holder:', ''))
        w_sport = clean_val(row.get('Sport', ''))
        score = float(row.get('_Match_Score', 0.0))
        print(
            f"  [{option_num}] Score {score:.3f} | {w_date} | {w_merchant} | {w_holder} | {w_sport}"
        )
    print("  [0] Skip (leave unmatched)")

    while True:
        choice = clean_val(input("Select candidate number to approve: "))
        if choice == "0":
            return None
        if choice.isdigit():
            choice_int = int(choice)
            if 1 <= choice_int <= len(candidates):
                return candidates.index[choice_int - 1]
        print("Invalid choice. Enter a number shown above.")


def is_likely_person_name(line):
    """Heuristic to identify cardholder name lines on statement headers."""
    s = clean_val(line)
    if not s:
        return False
    upper = s.upper()

    blocked_prefixes = (
        "ATTN", "PO BOX", "ACCOUNT", "STATEMENT", "TOTAL", "PAGE", "VISIT", "WWW",
        "REGIONS", "BILLING", "INDIVIDUAL", "MEMO", "MSU "
    )
    if upper.startswith(blocked_prefixes):
        return False

    if any(ch.isdigit() for ch in s):
        return False

    # Allow typical name characters only.
    if not re.match(r"^[A-Za-z .,'-]+$", s):
        return False

    # Person name should usually have at least 2 tokens (e.g., FIRST LAST).
    tokens = [t for t in re.split(r'\s+', s.strip()) if t]
    return len(tokens) >= 2


def is_likely_sport_or_group_name(line):
    """Heuristic to identify sport/group line near cardholder in statement header/address block."""
    s = clean_val(line)
    if not s:
        return False

    upper = s.upper()
    compact = upper.replace(" ", "")

    blocked_prefixes = (
        "ATTN", "PO BOX", "ACCOUNT", "STATEMENT", "TOTAL", "PAGE", "VISIT", "WWW",
        "REGIONS", "BILLING", "INDIVIDUAL", "MEMO", "FEB ", "MAR ", "JAN ", "APR ",
        "MAY ", "JUN ", "JUL ", "AUG ", "SEP ", "OCT ", "NOV ", "DEC ",
        "MISSISSIPPI STATE"
    )
    if upper.startswith(blocked_prefixes):
        return False

    blocked_exact = {
        "ACCOUNTNUMBER", "TOTALACTIVITY", "ACCOUNTINQUIRIES", "CARDHOLDERACTIVITY",
        "TRANSACTIONS", "DATE", "AMOUNT", "REFERENCENUMBER", "CATEGORY"
    }
    if compact in blocked_exact:
        return False

    if any(ch.isdigit() for ch in s):
        return False

    return bool(re.search(r'[A-Za-z]', s))


def is_header_group_candidate(line, card_holder=""):
    """Allow broader non-name header lines for sport/group detection near the address block."""
    s = clean_val(line)
    if not s:
        return False
    upper = s.upper()
    if upper == clean_val(card_holder).upper():
        return False
    if upper.startswith(("ATTN", "PO BOX", "MISSISSIPPI STATE", "ACCOUNT", "STATEMENT", "TOTAL", "PAGE")):
        return False
    if any(ch.isdigit() for ch in s):
        return False
    return True


def is_valid_sport_value(line, card_holder=""):
    """Validate final sport values and exclude boilerplate/header text."""
    s = clean_val(line)
    if not s or s.lower() == 'unknown':
        return False
    if s.upper() == clean_val(card_holder).upper():
        return False
    if s.upper().startswith(("VISIT ", "LOST OR STOLEN", "WWW.", "ACCOUNT ", "STATEMENT ")):
        return False
    return True


def extract_header_fields(first_page_text):
    """Extract sport, cardholder, and total activity from page 1 reliably."""
    lines = [clean_val(ln) for ln in (first_page_text or "").split('\n') if clean_val(ln)]

    sport = "Unknown"
    card_holder = "Unknown"
    total_activity = 0.0

    # Total activity usually appears adjacent to "TotalActivity" label.
    for i, line in enumerate(lines):
        if "TOTALACTIVITY" in line.upper().replace(" ", ""):
            for j in range(max(0, i - 3), min(len(lines), i + 3)):
                amt_match = re.search(r'\$[\d,]+\.\d{2}', lines[j])
                if amt_match:
                    total_activity = parse_amount(amt_match.group(0), default=0.0)
                    break
            if total_activity != 0.0:
                break

    # Find sport lines and pick the best nearby name (avoid "Attn Melissa Inmon").
    sport_indexes = [idx for idx, ln in enumerate(lines) if ln.upper().startswith("MSU ")]
    for idx in sport_indexes:
        sport_candidate = lines[idx]
        next_line = lines[idx + 1] if idx + 1 < len(lines) else ""
        prev_line = lines[idx - 1] if idx - 1 >= 0 else ""

        name_candidate = ""
        if is_likely_person_name(next_line):
            name_candidate = next_line
        elif is_likely_person_name(prev_line):
            name_candidate = prev_line

        if name_candidate:
            sport = sport_candidate
            card_holder = name_candidate
            break

    # Last fallback: first valid sport, and first likely uppercase name anywhere.
    if sport == "Unknown" and sport_indexes:
        sport = lines[sport_indexes[0]]

    if card_holder == "Unknown":
        for ln in lines:
            if is_likely_person_name(ln):
                card_holder = ln
                break

    return card_holder, sport, total_activity


def extract_last_name_from_filename(file_name):
    """Get expected cardholder last name from statement file name (e.g., Armistead-7824...)."""
    base = Path(file_name).stem
    # Remove trailing month/details in parentheses if present.
    base = re.sub(r'\s*\(.*\)\s*$', '', base)
    # Last name appears before first dash in current naming convention.
    last_name = base.split('-')[0].strip()
    return re.sub(r'[^A-Za-z\-\']', '', last_name).upper()


def extract_header_fields_with_filename(first_page_text, file_name):
    """Extract sport/cardholder/total, anchored by expected last name from file name."""
    lines = [clean_val(ln) for ln in (first_page_text or "").split('\n') if clean_val(ln)]
    expected_last = extract_last_name_from_filename(file_name)

    # Start with existing generic extraction.
    card_holder, sport, total_activity = extract_header_fields(first_page_text)

    matched_name_index = None

    # Stronger cardholder selection: choose name line containing expected last name.
    if expected_last:
        for idx, ln in enumerate(lines):
            ln_upper = ln.upper()
            if is_likely_person_name(ln) and expected_last in ln_upper:
                card_holder = ln
                matched_name_index = idx
                break

    # If we found a reliable name, resolve sport/group from lines around that name.
    if matched_name_index is not None:
        # Requirement: group is typically directly under parsed name in address block.
        below = lines[matched_name_index + 1] if matched_name_index + 1 < len(lines) else ""
        above = lines[matched_name_index - 1] if matched_name_index - 1 >= 0 else ""

        if is_header_group_candidate(below, card_holder):
            sport = below
        elif is_header_group_candidate(above, card_holder):
            sport = above
        else:
            # Nearby fallback around matched name.
            nearby = lines[max(0, matched_name_index - 4):min(len(lines), matched_name_index + 5)]
            nearby_group = next((s for s in nearby if is_header_group_candidate(s, card_holder)), "")
            if nearby_group:
                sport = nearby_group

    # Final guardrails against swapped values.
    if is_likely_sport_or_group_name(card_holder) and is_likely_person_name(sport):
        card_holder, sport = sport, card_holder

    # Ensure card holder aligns to statement file's last name requirement.
    if expected_last and expected_last not in clean_val(card_holder).upper():
        matching_name = next(
            (ln for ln in lines if is_likely_person_name(ln) and expected_last in ln.upper()),
            ""
        )
        if matching_name:
            card_holder = matching_name
        else:
            # Guaranteed filled fallback when OCR/text extraction is imperfect.
            card_holder = expected_last.title()

    # Ensure sport/group is never Unknown when statement text has any plausible group line.
    if sport == "Unknown" or not clean_val(sport):
        if matched_name_index is not None:
            nearby = lines[max(0, matched_name_index - 6):min(len(lines), matched_name_index + 7)]
            fallback_group = next((s for s in nearby if is_header_group_candidate(s, card_holder)), "")
            if fallback_group:
                sport = fallback_group

    if sport == "Unknown" or not clean_val(sport):
        fallback_group = next((s for s in lines if is_header_group_candidate(s, card_holder)), "")
        if fallback_group:
            sport = fallback_group

    # Never allow sport/group to equal the cardholder name.
    if clean_val(sport).upper() == clean_val(card_holder).upper():
        sport = "Unknown"

    return card_holder, sport, total_activity


def parse_transaction_line(line):
    """Parse one text line into (date, category, reference, transaction, amount, is_credit)."""
    if not line:
        return None

    compact = re.sub(r'\s+', ' ', str(line)).strip()
    # Match format: MM/DD MM/DD CCCC RRRRRRRRRRRRR MERCHANT... $$$
    # Date Tran Cat  Ref  Transaction        Amount
    date_match = re.match(
        r'^(\d{2}/\d{2})\s+(\d{2}/\d{2})\s+(\d{4})\s+(\d+)\s+(.+?)\s+(\$?[\d,]+\.\d{2})(?:\s*(CR))?\b',
        compact,
        flags=re.IGNORECASE,
    )
    if date_match:
        tran_date, _post_date, category, ref_num, transaction_desc, amount_str, cr_flag = date_match.groups()
        amount_str = amount_str.replace('$', '').replace(',', '')
        try:
            amount = float(amount_str)
        except ValueError:
            return None

        tail = compact[date_match.end():]
        is_credit = bool(cr_flag) or bool(re.search(r'\bCR\b', tail, flags=re.IGNORECASE)) or amount < 0
        if is_credit:
            amount = -abs(amount)

        return tran_date, category, ref_num, transaction_desc.strip(), amount, is_credit
    
    # Fallback to simpler date+amount pattern for text-only lines
    if not re.match(r'^\d{2}/\d{2}\b', compact):
        return None

    amount_matches = list(re.finditer(r'-?\$?\d[\d,]*\.\d{2}', compact))
    if not amount_matches:
        return None

    amount_match = amount_matches[-1]
    amount_str = amount_match.group(0).replace('$', '').replace(',', '')
    try:
        amount = float(amount_str)
    except ValueError:
        return None

    tail = compact[amount_match.end():]
    is_credit = bool(re.search(r'\bCR\b', tail, flags=re.IGNORECASE))
    if is_credit:
        amount = -abs(amount)

    transaction = compact[5:amount_match.start()].strip()
    transaction = re.sub(r'^\d{2}/\d{2}\s+', '', transaction).strip()
    if transaction == "":
        transaction = "Unknown Transaction"

    tx_date = compact[:5]
    return tx_date, "", "", transaction, amount, is_credit


def split_transaction_details(cell_text):
    """Split combined table cell text into category, reference number, and transaction.
    
    Expected format: [4 digits for category] [digits for reference number] [transaction description]
    """
    text = clean_val(cell_text)
    if not text:
        return "", "", ""

    # Prefer line-based parsing when PDF table extraction preserves line breaks.
    lines = [re.sub(r'\s+', ' ', ln).strip() for ln in text.split('\n') if clean_val(ln)]
    if len(lines) >= 3:
        category = lines[0]
        reference_number = lines[1]
        transaction = " ".join(lines[2:]).strip()
        return category, reference_number, transaction

    compact = re.sub(r'\s+', ' ', text).strip()

    # Extract first 4 digits as category
    category_match = re.match(r'^(\d{4})', compact)
    if category_match:
        category = category_match.group(1)
        rest = compact[category_match.end():].strip()
        
        # Extract next grouping of digits as reference number
        ref_match = re.match(r'^(\d+)', rest)
        if ref_match:
            reference_number = ref_match.group(1)
            transaction = rest[ref_match.end():].strip()
            return category, reference_number, transaction
        else:
            # No reference number found, rest is transaction
            return category, "", rest
    
    # If no 4-digit category found, return full text as transaction
    return "", "", compact

def _parse_pdf_bytes(pdf_bytes, file_name, all_tx, cycle_starts, cycle_ends):
    """Parse a single PDF's bytes and append its transactions to all_tx in-place."""
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        # Extract Card Holder, Sport, and Total from Page 1
        first_page_text = pdf.pages[0].extract_text() or ""
        card_holder, sport, total_activity = extract_header_fields_with_filename(first_page_text, file_name)
        cycle_start, cycle_end = extract_statement_cycle_range(first_page_text)
        if cycle_start and cycle_end:
            cycle_starts.append(cycle_start)
            cycle_ends.append(cycle_end)

        # Extract transactions from all pages after cover page.
        # Try table parsing first, then fallback to text parsing.
        for page in pdf.pages[1:]:
            parsed_count = 0
            table = page.extract_table()
            if table:
                for row in table:
                    row = [str(i).strip() if i else "" for i in row]
                    if len(row) >= 6 and re.match(r'\d{2}/\d{2}', row[0]):
                        amt = float(row[5].replace(',', '').replace('$', ''))
                        is_credit = any("CR" in str(c) for c in row)
                        if is_credit:
                            amt = -abs(amt)
                        category, reference_number, transaction = split_transaction_details(row[4])
                        transaction_text = transaction or row[4].replace('\n', ' ')
                        all_tx.append({
                            'Date': row[0],
                            'Category': category,
                            'Reference Number': reference_number,
                            'Transaction': transaction_text,
                            'Amount': amt,
                            'Card Holder': card_holder,
                            'Sport': sport,
                            'Total Activity': total_activity,
                            'Statement Credit': is_credit
                        })
                        parsed_count += 1

            if parsed_count == 0:
                page_text = page.extract_text() or ""
                for line in page_text.split('\n'):
                    parsed = parse_transaction_line(line)
                    if not parsed:
                        continue
                    tx_date, category, ref_num, transaction, amount, is_credit = parsed
                    all_tx.append({
                        'Date': tx_date,
                        'Category': category,
                        'Reference Number': ref_num,
                        'Transaction': transaction,
                        'Amount': amount,
                        'Card Holder': card_holder,
                        'Sport': sport,
                        'Total Activity': total_activity,
                        'Statement Credit': is_credit
                    })


def _finalize_bank_df(all_tx, cycle_starts, cycle_ends):
    bank_df = pd.DataFrame(all_tx)
    if cycle_starts and cycle_ends:
        bank_df.attrs['statement_cycle_start'] = min(cycle_starts)
        bank_df.attrs['statement_cycle_end'] = max(cycle_ends)
    return bank_df


def parse_bank_statements(zip_path):
    """Extracts all transactions from the ZIP of Regions Bank PDFs."""
    all_tx = []
    cycle_starts = []
    cycle_ends = []
    if hasattr(zip_path, 'seek'):
        zip_path.seek(0)
    with zipfile.ZipFile(zip_path, 'r') as z:
        for file_name in z.namelist():
            if file_name.lower().endswith('.pdf'):
                with z.open(file_name) as pdf_file:
                    _parse_pdf_bytes(pdf_file.read(), file_name, all_tx, cycle_starts, cycle_ends)
    return _finalize_bank_df(all_tx, cycle_starts, cycle_ends)


def parse_bank_statements_from_files(pdf_files):
    """Extracts all transactions from a list of (filename, bytes) PDF tuples.

    Args:
        pdf_files: list of (filename, bytes) — e.g. from Streamlit multi-file uploader.
    """
    all_tx = []
    cycle_starts = []
    cycle_ends = []
    for file_name, pdf_bytes in pdf_files:
        if file_name.lower().endswith('.pdf'):
            _parse_pdf_bytes(pdf_bytes, file_name, all_tx, cycle_starts, cycle_ends)
    return _finalize_bank_df(all_tx, cycle_starts, cycle_ends)

def run_procard_pipeline(
    workflow_csv,
    bank_zip=None,
    batch_date=None,
    manual_review_duplicates=True,
    prompt_for_missing_foapal=True,
    output_file=None,
    bank_pdf_files=None,
):
    if not batch_date:
        batch_date = datetime.now().strftime('%y%m%d')

    # 1. Load Workflow CSV
    workflow = pd.read_csv(workflow_csv, index_col=False)
    # Clean keys and values because of the tabs/quotes in your specific file
    workflow.columns = [c.replace("'", "").strip() for c in workflow.columns]
    for col in ['Merchant:', 'Date of Transaction:', 'Card Holder:']:
        workflow[col] = workflow[col].apply(clean_val)
    workflow['Amount: $'] = workflow['Amount: $'].apply(parse_amount)
    workflow['Amount:'] = workflow['Amount:'].apply(parse_amount)
    workflow['Statement Credit'] = False
    # Drop blank footer/header rows that have no date and no merchant (e.g. CSV metadata rows)
    workflow = workflow[~(workflow['Date of Transaction:'].eq('') & workflow['Merchant:'].eq(''))].reset_index(drop=True)

    # Build name->sport lookup from workflow for parser backfill.
    workflow_cardholder_col = first_matching_column(workflow, 'Card Holder:')
    workflow_sport_col = first_matching_column(workflow, 'Sport')

    workflow['_CardHolder_Norm'] = workflow_cardholder_col.apply(
        lambda x: re.sub(r'\s+', ' ', clean_val(x)).upper()
    )
    workflow['_Sport_Clean'] = workflow_sport_col.apply(clean_val)
    all_cardholder_names = set(workflow['_CardHolder_Norm'].dropna().astype(str).str.strip())
    name_to_sport = {}
    for name, grp in workflow.groupby('_CardHolder_Norm', dropna=True):
        if not clean_val(name):
            continue
        sport_candidates = [
            clean_val(s) for s in grp['_Sport_Clean'].tolist()
            if clean_val(s) and re.sub(r'\s+', ' ', clean_val(s)).upper() not in all_cardholder_names
        ]
        if sport_candidates:
            # Use most common non-name sport value for this cardholder.
            name_to_sport[name] = pd.Series(sport_candidates).value_counts().idxmax()

    # 2. Parse Bank Statements
    print("Parsing Bank Statements...")
    if bank_pdf_files is not None:
        bank_df = parse_bank_statements_from_files(bank_pdf_files)
    elif bank_zip is not None:
        bank_df = parse_bank_statements(bank_zip)
    else:
        raise ValueError("Either bank_zip or bank_pdf_files must be provided.")
    statement_cycle_start = bank_df.attrs.get('statement_cycle_start') if hasattr(bank_df, 'attrs') else None
    statement_cycle_end = bank_df.attrs.get('statement_cycle_end') if hasattr(bank_df, 'attrs') else None
    if statement_cycle_start and statement_cycle_end:
        print(f"Detected statement cycle from PDFs: {format_statement_cycle(statement_cycle_start, statement_cycle_end).replace(' - ', ' to ')}")

    # Ensure sport is always populated: backfill from workflow by cardholder name.
    if not bank_df.empty:
        bank_df['Card Holder'] = bank_df['Card Holder'].apply(clean_val)
        bank_df['Sport'] = bank_df['Sport'].apply(clean_val)
        missing_sport_mask = bank_df.apply(
            lambda row: not is_valid_sport_value(row.get('Sport', ''), row.get('Card Holder', '')),
            axis=1
        )
        if missing_sport_mask.any():
            bank_df['_CardHolder_Norm'] = bank_df['Card Holder'].apply(
                lambda x: re.sub(r'\s+', ' ', clean_val(x)).upper()
            )
            bank_df.loc[missing_sport_mask, 'Sport'] = bank_df.loc[missing_sport_mask, '_CardHolder_Norm'].map(name_to_sport).fillna(bank_df.loc[missing_sport_mask, 'Sport'])
            bank_df.drop(columns=['_CardHolder_Norm'], inplace=True)

    # Transaction counts
    bank_transaction_count = len(bank_df)
    workflow_transaction_count = int(
        workflow['Date of Transaction:'].apply(lambda x: normalize_mmdd(x) != "").sum()
    )
    transaction_count_difference = bank_transaction_count - workflow_transaction_count

    print(f"Bank statement transactions found: {bank_transaction_count}")
    print(f"Workflow CSV transactions found: {workflow_transaction_count}")
    print(f"Difference (bank - workflow): {transaction_count_difference}")

    # 3. VALIDATION: Check if every bank transaction exists in the Workflow CSV
    # Match criteria: Date (MM/DD), Amount, and Transaction text
    workflow['_Date_MMDD'] = workflow['Date of Transaction:'].apply(normalize_mmdd)
    missing_transactions = []
    matched_workflow_rows = set()
    cardholder_sport_votes = {}
    reused_workflow_match_count = 0
    manual_approved_match_count = 0
    for _, tx in bank_df.iterrows():
        tx_date = normalize_mmdd(tx.get('Date', ''))
        tx_amount = round(parse_amount(tx.get('Amount', 0.0)), 2)
        tx_transaction = clean_val(tx.get('Transaction', ''))

        unmatched_match, match_score = find_best_workflow_match(
            workflow, tx_date, tx_amount, tx_transaction, matched_workflow_rows
        )

        # If unique row is unavailable, allow high-confidence reuse of an already matched
        # workflow row (same date/amount + very strong merchant similarity).
        reused_match = False
        if unmatched_match.empty:
            reused_candidate, reused_score = find_best_workflow_match(
                workflow, tx_date, tx_amount, tx_transaction, set(), threshold=0.90
            )
            if not reused_candidate.empty:
                unmatched_match = reused_candidate
                match_score = reused_score
                reused_match = True

        # Final fallback: amount + strong merchant similarity (date drift tolerant).
        if unmatched_match.empty:
            relaxed_candidate, relaxed_score = find_best_workflow_match_relaxed(
                workflow, tx_date, tx_amount, tx_transaction, set(), threshold=0.93
            )
            if not relaxed_candidate.empty:
                unmatched_match = relaxed_candidate
                match_score = relaxed_score
                reused_match = True

        # Manual review for possible duplicates/alternates before declaring unmatched.
        if unmatched_match.empty and manual_review_duplicates:
            review_candidates = get_workflow_match_candidates(
                workflow,
                tx_date,
                tx_amount,
                tx_transaction,
                matched_workflow_rows,
                threshold=0.55,
                relaxed=False,
                max_candidates=8,
            )
            if review_candidates.empty:
                review_candidates = get_workflow_match_candidates(
                    workflow,
                    tx_date,
                    tx_amount,
                    tx_transaction,
                    set(),
                    threshold=0.88,
                    relaxed=True,
                    max_candidates=8,
                )

            if not review_candidates.empty:
                approved_index = prompt_user_to_approve_match(tx, review_candidates)
                if approved_index is not None:
                    unmatched_match = workflow.loc[[approved_index]].copy()
                    reused_match = approved_index in matched_workflow_rows
                    manual_approved_match_count += 1

        if unmatched_match.empty:
            missing_transactions.append(tx)
            continue

        matched_index = unmatched_match.index[0]
        if not reused_match:
            matched_workflow_rows.add(matched_index)
        else:
            reused_workflow_match_count += 1
        workflow.at[matched_index, 'Statement Credit'] = bool(tx.get('Statement Credit', False))

        matched_sport = clean_val(workflow.at[matched_index, 'Sport']) if 'Sport' in workflow.columns else ""
        matched_cardholder = clean_val(tx.get('Card Holder', ''))
        if matched_cardholder and matched_sport:
            cardholder_sport_votes.setdefault(matched_cardholder, []).append(matched_sport)

    # Final sport backfill from successfully matched workflow rows for each cardholder.
    print(f"{len(matched_workflow_rows)} PDF transactions matched to workflow rows.")
    if not bank_df.empty and cardholder_sport_votes:
        def resolve_cardholder_sport(row):
            current_sport = clean_val(row.get('Sport', ''))
            if is_valid_sport_value(current_sport, row.get('Card Holder', '')):
                return current_sport
            holder = clean_val(row.get('Card Holder', ''))
            votes = cardholder_sport_votes.get(holder, [])
            if not votes:
                return current_sport
            return pd.Series(votes).value_counts().idxmax()

        bank_df['Sport'] = bank_df.apply(resolve_cardholder_sport, axis=1)

    # Reverse reconciliation: workflow rows that did not match any PDF transaction.
    workflow_transaction_mask = workflow['Date of Transaction:'].apply(lambda x: normalize_mmdd(x) != "")
    workflow_unmatched_indices = [
        idx for idx in workflow[workflow_transaction_mask].index
        if idx not in matched_workflow_rows
    ]
    workflow_only_rows = workflow.loc[workflow_unmatched_indices, [
        'Date of Transaction:', 'Merchant:', 'Amount: $', 'Card Holder:', 'Sport'
    ]].copy()
    workflow_only_rows.columns = ['Date', 'Merchant', 'Amount', 'Card Holder', 'Sport']
    workflow_only_rows['Warning Missing'] = ""
    previous_statement_indices = set()
    if statement_cycle_start and statement_cycle_end and not workflow_only_rows.empty:
        parsed_dates = workflow_only_rows['Date'].apply(
            lambda d: parse_workflow_date(d, statement_cycle_start, statement_cycle_end)
        )
        workflow_only_rows['Warning Missing'] = parsed_dates.apply(
            lambda d: classify_date_outside_cycle(d, statement_cycle_start, statement_cycle_end)
        )
        previous_statement_indices = set(
            workflow_only_rows.index[
                workflow_only_rows['Warning Missing'].fillna('').astype(str).str.startswith('Likely on previous statement')
            ]
        )

    if missing_transactions:
        print("\n!!! ALERT: The following bank statement transactions were NOT found in the Workflow CSV:")
        for m in missing_transactions:
            tx_date = normalize_mmdd(m.get('Date', ''))
            tx_amount = round(parse_amount(m.get('Amount', 0.0)), 2)
            tx_transaction = clean_val(m.get('Transaction', ''))

            partial_match, match_score = find_best_workflow_match(
                workflow, tx_date, tx_amount, tx_transaction, set(), threshold=0.55
            )
            
            sport = ""
            if not partial_match.empty:
                sport = clean_val(partial_match.iloc[0].get('Sport', ''))
            
            sport_str = f" | Sport: {sport}" if sport else ""
            print(f"  - {m['Date']} | {m['Transaction']} | ${m['Amount']} ({m['Card Holder']}){sport_str}")
        print("These will be excluded from the final file feed.")
    else:
        print("Verification Successful: All bank transactions accounted for.")

    if not workflow_only_rows.empty:
        previous_mask = workflow_only_rows['Warning Missing'].fillna('').astype(str).str.startswith('Likely on previous statement')
        next_mask = workflow_only_rows['Warning Missing'].fillna('').astype(str).str.startswith('Likely on next statement')

        if previous_mask.all():
            cycle_str = ""
            if statement_cycle_start and statement_cycle_end:
                cycle_str = f" (statement cycle {format_statement_cycle(statement_cycle_start, statement_cycle_end)})"
            print(
                f"\nSummary: All {len(workflow_only_rows)} workflow rows not found in PDFs are likely on the previous month's statement{cycle_str}."
            )
        else:
            previous_count = int(previous_mask.sum())
            next_count = int(next_mask.sum())
            other_count = len(workflow_only_rows) - previous_count - next_count
            print(
                f"\nSummary: Workflow rows not found in PDFs = {len(workflow_only_rows)} "
                f"(likely previous statement: {previous_count}, likely next statement: {next_count}, other: {other_count})."
            )
    else:
        print("Workflow verification successful: All workflow transactions were found in bank statements.")

    if reused_workflow_match_count:
        print(f"High-confidence duplicate matches used: {reused_workflow_match_count}")
    if manual_approved_match_count:
        print(f"Manual duplicate/alternate matches approved: {manual_approved_match_count}")

    workflow_transaction_rows = workflow.loc[workflow_transaction_mask].copy()
    comparable_workflow_rows = workflow_transaction_rows.loc[
        workflow_transaction_rows.index.isin(matched_workflow_rows)
    ].copy()
    pdf_parsed_total = float(round(bank_df['Amount'].apply(parse_amount).sum(), 2)) if not bank_df.empty else 0.0
    comparable_workflow_total = float(round(comparable_workflow_rows['Amount: $'].apply(parse_amount).sum(), 2))
    excluded_previous_total = float(round(
        workflow_transaction_rows.loc[
            workflow_transaction_rows.index.isin(previous_statement_indices), 'Amount: $'
        ].apply(parse_amount).sum(),
        2,
    ))
    total_difference = float(round(pdf_parsed_total - comparable_workflow_total, 2))

    print("\nTransaction Total Comparison:")
    print(f"  PDF parsed statement total: ${pdf_parsed_total:,.2f}")
    print(f"  Comparable workflow total: ${comparable_workflow_total:,.2f}")
    if previous_statement_indices:
        print(
            f"  Excluded likely previous-statement workflow total: ${excluded_previous_total:,.2f} "
            f"across {len(previous_statement_indices)} row(s)"
        )
    print(f"  Difference (PDF - workflow): ${total_difference:,.2f}")

    # 4. STAGE 2: CREATE TEAMS WORKFLOW (Cleaned version of CSV)
    # Keep only workflow rows that were matched to parsed PDF transactions,
    # then expand split transactions so the final output reflects PDF∩workflow rows only.
    matched_pdf_mask = workflow.index.to_series().isin(matched_workflow_rows)
    removed_count = int((~matched_pdf_mask).sum())
    workflow_kept = workflow.loc[matched_pdf_mask].copy()
    if removed_count:
        print(
            f"Excluded {removed_count} workflow row(s) not matched to parsed PDF transactions "
            f"from TEAMS WORKFLOW/FILE FEED."
        )

    # Expand split transactions on the filtered rows.
    # Each row with multiple FOAPAL lines becomes one row per line (main + splits).
    workflow_kept_before = len(workflow_kept)
    workflow_kept = expand_split_transactions(workflow_kept)
    split_rows_added = len(workflow_kept) - workflow_kept_before
    print(f"Split transaction expansion: {workflow_kept_before} rows -> {len(workflow_kept)} rows ({split_rows_added} split rows added).")

    teams_workflow = pd.DataFrame({
        'Current Task Name': first_matching_column(workflow_kept, 'Current Task Name'),
        'Sport': first_matching_column(workflow_kept, 'Sport'),
        'Merchant:': first_matching_column(workflow_kept, 'Merchant:'),
        'Date of Transaction:': first_matching_column(workflow_kept, 'Date of Transaction:'),
        'Amount: $': first_matching_column(workflow_kept, 'Amount: $'),
        'Card Holder:': first_matching_column(workflow_kept, 'Card Holder:'),
        'Fund Code:': first_matching_column(workflow_kept, 'Fund Code:'),
        'Organization:': first_matching_column(workflow_kept, 'Organization:'),
        'Account:': first_matching_column(workflow_kept, 'Account:'),
        'Program:': first_matching_column(workflow_kept, 'Program:'),
        'AD Code:': first_matching_column(workflow_kept, 'AD Code:'),
        'Amount:': first_matching_column(workflow_kept, 'Amount:'),
        'Statement Credit': workflow_kept['Statement Credit']
    }).copy()

    # Prompt user for missing FOAPAL values and fill before file feed generation
    if prompt_for_missing_foapal:
        teams_workflow = fill_missing_foapal(teams_workflow)

    # Append PROCARD net total row at the bottom of TEAMS WORKFLOW.
    teams_workflow = append_procard_total_row(teams_workflow, total_amount=comparable_workflow_total)

    # 5. STAGE 3: TRANSFORM TO FILE FEED
    file_feed = build_file_feed_from_teams_workflow(teams_workflow, batch_date)

    # 6. OUTPUT TO XLSX
    # Build a clean view of the original workflow for comparison (key columns only)
    workflow_review = workflow[[
        'Date of Transaction:', 'Merchant:', 'Amount: $', 'Card Holder:'
    ]].copy()
    workflow_review.columns = ['Date', 'Merchant', 'Amount', 'Card Holder']

    # Build a clean view of parsed PDF transactions for comparison
    bank_review_columns = ['Date', 'Category', 'Reference Number', 'Transaction', 'Amount', 'Card Holder', 'Sport', 'Total Activity']
    bank_review = bank_df[bank_review_columns].copy() if not bank_df.empty else pd.DataFrame(columns=bank_review_columns)

    output_file = output_file or 'Final_ProCard_Upload.xlsx'
    try:
        write_output_workbook(output_file, file_feed, teams_workflow, bank_review, workflow_review, workflow_only_rows)
    except PermissionError:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'Final_ProCard_Upload_{timestamp}.xlsx'
        write_output_workbook(output_file, file_feed, teams_workflow, bank_review, workflow_review, workflow_only_rows)
        print(f"\nFile was locked — saved as: {output_file}")

    print(f"\nFinal processing complete. File saved: {output_file}")
    statement_cycle = format_statement_cycle(statement_cycle_start, statement_cycle_end)
    return {
        'transaction_count_difference': transaction_count_difference,
        'pdf_parsed_total': pdf_parsed_total,
        'comparable_workflow_total': comparable_workflow_total,
        'excluded_previous_statement_total': excluded_previous_total,
        'total_difference': total_difference,
        'statement_cycle': statement_cycle,
        'output_file': output_file,
        'rows_after_split': len(workflow_kept),
        'workflow_matched_count': len(matched_workflow_rows),
        'workflow_previous_cycle_count': len(previous_statement_indices),
    }

def parse_cli_args():
    """Parse command-line inputs so the script can run with any month's files."""
    parser = argparse.ArgumentParser(
        description="Reconcile ProCard workflow CSV with bank statement PDFs in a ZIP file."
    )
    parser.add_argument(
        "workflow_csv",
        help="Path to the ProCard workflow CSV file."
    )
    parser.add_argument(
        "bank_zip",
        help="Path to the ZIP containing bank statement PDFs."
    )
    parser.add_argument(
        "--batch-date",
        dest="batch_date",
        default=None,
        help="Batch date for FILE FEED Col6 in YYMMDD format (default: today's date)."
    )
    parser.add_argument(
        "--no-manual-review",
        action="store_true",
        help="Disable interactive duplicate/alternate match review prompts."
    )
    parser.add_argument(
        "--no-foapal-prompt",
        action="store_true",
        help="Disable interactive FOAPAL input prompts."
    )
    return parser.parse_args()


if __name__ == '__main__':
    args = parse_cli_args()
    result = run_procard_pipeline(
        workflow_csv=args.workflow_csv,
        bank_zip=args.bank_zip,
        batch_date=args.batch_date,
        manual_review_duplicates=not args.no_manual_review,
        prompt_for_missing_foapal=not args.no_foapal_prompt,
    )
    print(f"\nReturned summary: {result}")