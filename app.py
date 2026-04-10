import io
import os
import re
import tempfile
import textwrap
from collections import Counter
from contextlib import redirect_stdout
from datetime import datetime
from calendar import monthrange

import pandas as pd
import streamlit as st
from openpyxl.styles import Font

from processor import FILE_FEED_COLUMN_WIDTHS, build_file_feed_from_teams_workflow, run_procard_pipeline


st.set_page_config(page_title="ProCard Reconciliation App", layout="wide")

# ── Brand Primary  ────────────────────────────────────
BRAND_PRIME = "#5D1725"
BRAND_GREY = "#777777"
BRAND_LIGHT = "#C1C6C8"

# ── Brand Accent  ─────────────────────────────────────
BRAND_ACCENT = "#A69F88"
BRAND_GREEN = "#8F993E"
BRAND_ORANGE = "#A9431E"
BRAND_GOLD = "#C99700"
BRAND_CREAM = "#DAC79D"

WHITE = "#FFFFFF"
BLACK = "#000000"

STEP_TITLES = [
    "Welcome",
    "Upload Files",
    "Parse & Match",
    "FOAPAL Review",
    "Download Output",
]

WORKBOOK_FONT_NAME = "Aptos Narrow"


def apply_workbook_font(workbook, font_name=WORKBOOK_FONT_NAME, font_size=11):
    """Apply default workbook font (including Normal style) across all sheets."""
    font = Font(name=font_name, size=font_size)

    # Set workbook default "Normal" style so Excel width rendering is consistent.
    try:
        for named_style in getattr(workbook, "_named_styles", []):
            if getattr(named_style, "name", "") == "Normal":
                named_style.font = font
                break
    except Exception:
        pass

    for ws in workbook.worksheets:
        if ws.max_row <= 0 or ws.max_column <= 0:
            continue
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = font


def hex_to_rgba(hex_color, alpha):
    h = hex_color.lstrip("#")
    if len(h) != 6:
        return f"rgba(0,0,0,{alpha})"
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return f"rgba({r},{g},{b},{alpha})"


def render_helper_box(message, tone="info", target=None, compact=False):
    color_map = {
        "info": BRAND_GOLD,
        "warning": BRAND_GOLD,
        "success": BRAND_GREEN,
        "error": BRAND_ORANGE,
        "prime": BRAND_PRIME,
    }
    base = color_map.get(tone, BRAND_GOLD)
    bg = hex_to_rgba(base, 0.14)
    border = hex_to_rgba(base, 0.45)
    host = target if target is not None else st
    if compact:
        html = (
            f"<div style='text-align:center;'>"
            f"<div style='display:inline-block; border-radius:8px; border:1px solid {border}; "
            f"background:{bg}; padding:0.65rem 0.85rem; color:{BLACK};'>{message}</div>"
            f"</div>"
        )
    else:
        html = (
            f"<div style='border-radius:8px; border:1px solid {border}; "
            f"background:{bg}; padding:0.65rem 0.85rem; color:{BLACK};'>{message}</div>"
        )
    host.markdown(
        html,
        unsafe_allow_html=True,
    )


if hasattr(st, "dialog"):
    @st.dialog("Missing FOAPAL Values", width="medium")
    def render_foapal_missing_confirm_dialog(missing_count):
        st.error(
            f"{missing_count} row(s) are still missing FOAPAL values. Do you want to proceed to Download Output with those values still blank?"
        )
        confirm_col1, confirm_spacer, confirm_col2 = st.columns([2.2, 5.6, 2.6])
        with confirm_col1:
            if st.button("Continue Review", key="foapal_review_more"):
                st.session_state.foapal_next_confirm = False
                st.rerun()
        with confirm_col2:
            if st.button("Proceed to Download", key="foapal_confirm_next", use_container_width=True):
                st.session_state.foapal_next_confirm = False
                go_next()
                st.rerun()
else:
    def render_foapal_missing_confirm_dialog(missing_count):
        render_helper_box(
            f"{missing_count} row(s) are still missing FOAPAL values. Do you want to proceed to Download Output with those values still blank?",
            "error",
        )
        st.markdown("<div style='height: 0.35rem;'></div>", unsafe_allow_html=True)
        confirm_col1, confirm_spacer, confirm_col2 = st.columns([1.6, 7.4, 1.8])
        with confirm_col1:
            if st.button("Review More", key="foapal_review_more"):
                st.session_state.foapal_next_confirm = False
                st.rerun()
        with confirm_col2:
            if st.button("Proceed Anyway", key="foapal_confirm_next"):
                st.session_state.foapal_next_confirm = False
                go_next()
                st.rerun()


def apply_brand_theme():
    st.markdown(
        f"""
        <style>
        :root {{
            --brand-prime: {BRAND_PRIME};
            --brand-grey: {BRAND_GREY};
            --brand-light: {BRAND_LIGHT};
            --brand-accent: {BRAND_ACCENT};
            --brand-green: {BRAND_GREEN};
            --brand-orange: {BRAND_ORANGE};
            --brand-gold: {BRAND_GOLD};
            --brand-cream: {BRAND_CREAM};
            --white: {WHITE};
            --black: {BLACK};
        }}

        .stApp {{
            background: var(--white);
            color: var(--black);
        }}

        h1, h2, h3 {{
            color: var(--brand-prime);
        }}

        [data-testid="stSidebar"] {{
            background: var(--brand-light);
        }}

        [data-testid="stMetric"] {{
            background: var(--white);
            border: 1px solid var(--brand-light);
            border-radius: 10px;
            padding: 8px;
        }}

        [data-testid="stAlert"] {{
            border-radius: 8px;
            border: 1px solid var(--brand-accent);
        }}

        div.stButton > button {{
            background: var(--brand-prime);
            color: var(--white);
            border: 1px solid var(--brand-prime);
            border-radius: 8px;
        }}

        div.stButton > button:hover {{
            background: var(--brand-orange);
            border-color: var(--brand-orange);
            color: var(--white);
        }}

        div.stButton > button p,
        div.stFormSubmitButton > button p,
        div.stDownloadButton > button p {{
            white-space: nowrap !important;
            overflow: visible !important;
            margin: 0 !important;
        }}

        div.stDownloadButton > button {{
            background: var(--brand-prime);
            color: var(--white);
            border: 1px solid var(--brand-prime);
            border-radius: 8px;
        }}

        div.stFormSubmitButton > button {{
            background: var(--brand-prime) !important;
            color: var(--white) !important;
            border: 1px solid var(--brand-prime) !important;
            border-radius: 10px !important;
            font-weight: 700 !important;
            box-shadow: 0 2px 8px rgba(93, 23, 37, 0.28) !important;
        }}

        div.stFormSubmitButton > button:hover {{
            background: var(--brand-orange) !important;
            border-color: var(--brand-orange) !important;
            color: var(--white) !important;
        }}

        /* Radio selector (filled circle) */
        [data-testid="stRadio"] input[type="radio"] {{
            accent-color: var(--brand-prime) !important;
        }}

        [data-testid="stRadio"] label,
        [data-testid="stRadio"] p {{
            color: var(--black) !important;
        }}

        [data-testid="stProgressBar"] div[role="progressbar"] {{
            background-color: var(--brand-prime) !important;
        }}

        [data-testid="stProgressBar"] > div > div > div {{
            background-color: var(--brand-prime) !important;
        }}

        [data-testid="stProgressBar"] div[role="progressbar"] > div {{
            background-color: var(--brand-prime) !important;
        }}

        .stTextInput input,
        .stSelectbox div[data-baseweb="select"] > div,
        .stFileUploader,
        .stDataFrame {{
            border-color: var(--brand-light) !important;
        }}

        .st-key-mode_box,
        .st-key-upload_mode_box,
        .st-key-workflow_upload_box,
        .st-key-bank_upload_box,
        .st-key-batch_date_box,
        .st-key-resume_upload_box,
        .st-key-resume_batch_date_box {{
            border: 1px solid rgba(93, 23, 37, 0.24) !important;
            background-color: rgba(93, 23, 37, 0.03) !important;
            border-radius: 10px !important;
            padding: 0.4rem 0.7rem 0.2rem 0.7rem !important;
        }}

        /* Row-nav buttons: ghost outline to distinguish from solid page-nav buttons */
        .st-key-foapal_prev_row button,
        .st-key-foapal_next_row button {{
            background: transparent !important;
            color: var(--brand-prime) !important;
            border: 1.5px solid var(--brand-prime) !important;
            box-shadow: none !important;
        }}
        .st-key-foapal_prev_row button:hover,
        .st-key-foapal_next_row button:hover {{
            background: rgba(93, 23, 37, 0.08) !important;
            color: var(--brand-prime) !important;
            border-color: var(--brand-prime) !important;
        }}

        .st-key-foapal_prev_row_submit button,
        .st-key-foapal_next_row_submit button {{
            background: transparent !important;
            color: var(--brand-prime) !important;
            border: 1.5px solid var(--brand-prime) !important;
            box-shadow: none !important;
        }}
        .st-key-foapal_prev_row_submit button:hover,
        .st-key-foapal_next_row_submit button:hover {{
            background: rgba(93, 23, 37, 0.08) !important;
            color: var(--brand-prime) !important;
            border-color: var(--brand-prime) !important;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_top_logo():
    logo_png_path = os.path.join(os.path.dirname(__file__), "logo.png")
    logo_pdf_path = os.path.join(os.path.dirname(__file__), "ScriptState_Maroon(RGB).pdf")
    if os.path.exists(logo_png_path):
        st.image(logo_png_path, width=230)
        return

    if not os.path.exists(logo_pdf_path):
        return

    # Preferred: render first page of PDF to PNG if PyMuPDF is available.
    try:
        import fitz  # type: ignore

        with fitz.open(logo_pdf_path) as doc:
            if doc.page_count > 0:
                page = doc.load_page(0)
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
                st.image(pix.tobytes("png"), width=230)
                return
    except Exception:
        pass


def init_state():
    defaults = {
        "step": 0,
        "workflow_file": None,
        "zip_file": None,
        "pdf_files": [],
        "input_mode": "ZIP file",
        "session_mode": "New reconciliation",
        "resume_file": None,
        "resume_requested": False,
        "parse_requested": False,
        "parse_running": False,
        "batch_date": default_batch_date(),
        "processing_done": False,
        "processing_log": "",
        "summary": {},
        "teams_workflow": None,
        "bank_review": None,
        "workflow_review": None,
        "workflow_only_rows": None,
        "source_output_file": "",
        "foapal_review_df": None,
        "foapal_current_pos": 0,
        "foapal_next_confirm": False,
        "foapal_review_skipped": False,
        "teams_workflow_original": None,
        "final_output_bytes": None,
        "final_output_name": "",
        "final_output_signature": None,
    }
    for key, val in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = val


def go_next():
    st.session_state.step = min(st.session_state.step + 1, len(STEP_TITLES) - 1)


def request_resume_load():
    st.session_state.resume_requested = True


def request_parse_run():
    st.session_state.parse_requested = True


def default_batch_date():
    today = datetime.now()
    last_day = monthrange(today.year, today.month)[1]
    return today.replace(day=last_day).strftime("%y%m%d")


def is_last_day_batch_date_YYMMDD(value):
    try:
        parsed = datetime.strptime(str(value).strip(), "%y%m%d")
    except Exception:
        return False
    return parsed.day == monthrange(parsed.year, parsed.month)[1]


def reset_upload_step_state():
    st.session_state.workflow_file = None
    st.session_state.zip_file = None
    st.session_state.pdf_files = []
    st.session_state.input_mode = "ZIP file"
    st.session_state.session_mode = "New reconciliation"
    st.session_state.resume_file = None
    st.session_state.batch_date = default_batch_date()
    if "workflow_upload" in st.session_state:
        del st.session_state["workflow_upload"]
    if "zip_upload" in st.session_state:
        del st.session_state["zip_upload"]
    if "pdf_uploads" in st.session_state:
        del st.session_state["pdf_uploads"]


def reset_parse_step_state():
    st.session_state.processing_done = False
    st.session_state.parse_requested = False
    st.session_state.parse_running = False
    st.session_state.processing_log = ""
    st.session_state.summary = {}
    st.session_state.teams_workflow = None
    st.session_state.teams_workflow_original = None
    st.session_state.bank_review = None
    st.session_state.workflow_review = None
    st.session_state.workflow_only_rows = None
    st.session_state.source_output_file = ""
    st.session_state.foapal_review_df = None
    st.session_state.foapal_current_pos = 0
    st.session_state.foapal_next_confirm = False
    st.session_state.foapal_review_skipped = False
    st.session_state.final_output_bytes = None
    st.session_state.final_output_name = ""
    st.session_state.final_output_signature = None


def reset_foapal_step_state():
    if st.session_state.teams_workflow_original is not None:
        st.session_state.teams_workflow = st.session_state.teams_workflow_original.copy(deep=True)
    st.session_state.foapal_current_pos = 0
    st.session_state.foapal_next_confirm = False
    st.session_state.foapal_review_skipped = False
    st.session_state.final_output_bytes = None
    st.session_state.final_output_name = ""
    st.session_state.final_output_signature = None


def go_back():
    current_step = st.session_state.step
    if current_step == 1:
        reset_upload_step_state()
    elif current_step == 2:
        reset_parse_step_state()
    elif current_step == 3:
        reset_foapal_step_state()
    st.session_state.step = max(st.session_state.step - 1, 0)


def is_blank(val):
    return pd.isna(val) or str(val).strip() == "" or str(val).strip().lower() == "nan"


def parse_amount_ui(val):
    if is_blank(val):
        return 0.0
    s = str(val).replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def normalize_date_mmdd_ui(val):
    if is_blank(val):
        return ""

    # Handle pandas/Excel datetimes first.
    try:
        dt = pd.to_datetime(val, errors="coerce")
        if pd.notna(dt):
            return dt.strftime("%m/%d")
    except Exception:
        pass

    s = str(val).strip()
    if not s:
        return ""

    # Fallback for text dates already in MM/DD or MM/DD/YYYY form.
    m = re.search(r'(\d{1,2})/(\d{1,2})(?:/\d{2,4})?', s)
    if m:
        return f"{int(m.group(1)):02d}/{int(m.group(2)):02d}"

    # Fallback for ISO-like strings that may survive as text.
    m_iso = re.search(r'(\d{4})-(\d{1,2})-(\d{1,2})', s)
    if m_iso:
        return f"{int(m_iso.group(2)):02d}/{int(m_iso.group(3)):02d}"

    return ""


def normalize_merchant_ui(val):
    s = "" if is_blank(val) else str(val).upper()
    return re.sub(r'[^A-Z0-9]+', ' ', s).strip()


def build_row_key(date_val, amount_val, merchant_val):
    return (
        normalize_date_mmdd_ui(date_val),
        round(parse_amount_ui(amount_val), 2),
        normalize_merchant_ui(merchant_val),
    )


def build_row_key_amount_date(date_val, amount_val):
    """Relaxed reconciliation key for UI diagnostics: date + amount only."""
    return (
        normalize_date_mmdd_ui(date_val),
        round(parse_amount_ui(amount_val), 2),
    )


def select_rows_by_key_counts(df, key_col, counter_map):
    if df.empty or not counter_map:
        return pd.DataFrame(columns=df.columns)
    chunks = []
    for key, count in counter_map.items():
        if count <= 0:
            continue
        chunks.append(df.loc[df[key_col] == key].head(int(count)))
    return pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame(columns=df.columns)


def missing_required_mask(df, required_cols):
    return df[required_cols].apply(lambda col: col.map(is_blank)).any(axis=1)


def sanitize_for_output(df):
    """Remove nulls from output sheets to prevent blank/null export cells."""
    cleaned = df.copy()
    return cleaned.fillna("")


def normalize_foapal_code(val):
    s = "" if is_blank(val) else str(val).strip()
    if re.fullmatch(r"\d+\.0+", s):
        return s.split(".")[0]
    return s


def normalize_editable_foapal(df, required_cols):
    """Keep FOAPAL editor fields as text so Streamlit doesn't coerce them."""
    normalized = df.copy()
    for col in required_cols:
        normalized[col] = normalized[col].map(normalize_foapal_code)
    return normalized


def foapal_sort_key(value):
    return (0, int(value)) if value.isdigit() else (1, value.lower())


def build_foapal_dropdown_options(df, required_cols):
    options_by_col = {}
    value_counts_by_col = {}
    for col in required_cols:
        values = [normalize_foapal_code(val) for val in df[col] if not is_blank(val)]
        counts = pd.Series(values).value_counts() if values else pd.Series(dtype="int64")
        value_counts_by_col[col] = counts
        options_by_col[col] = sorted(set(values), key=foapal_sort_key)

    # Resolve Account/Program overlaps by keeping each code in the column where it appears more often.
    if 'Account:' in options_by_col and 'Program:' in options_by_col:
        account_set = set(options_by_col['Account:'])
        program_set = set(options_by_col['Program:'])
        overlap = account_set & program_set
        for code in overlap:
            account_count = int(value_counts_by_col['Account:'].get(code, 0))
            program_count = int(value_counts_by_col['Program:'].get(code, 0))
            if account_count >= program_count:
                if code in options_by_col['Program:']:
                    options_by_col['Program:'].remove(code)
            else:
                if code in options_by_col['Account:']:
                    options_by_col['Account:'].remove(code)

    for col in required_cols:
        options_by_col[col] = [""] + options_by_col[col]
    return options_by_col


def selectbox_value(label, current_value, options, allow_outside_current=True):
    current = normalize_foapal_code(current_value)
    choices = list(options)
    if allow_outside_current and current and current not in choices:
        choices.append(current)
    selected_index = choices.index(current) if current in choices else 0
    try:
        return st.selectbox(
            label,
            options=choices,
            index=selected_index,
            accept_new_options=True,
        )
    except TypeError:
        # Fallback for older Streamlit versions that don't support accept_new_options.
        return st.selectbox(label, options=choices, index=selected_index)


def clamp_review_position(missing_ids):
    if not missing_ids:
        st.session_state.foapal_current_pos = 0
        return
    st.session_state.foapal_current_pos = min(
        st.session_state.foapal_current_pos,
        len(missing_ids) - 1,
    )


def render_step_nav():
    completed_steps = st.session_state.step
    last_idx = max(1, len(STEP_TITLES) - 1)
    progress = completed_steps / last_idx
    progress_pct = int(round(progress * 100))
    fill_width_css = "0%" if progress_pct == 0 else f"min(100%, calc({progress_pct}% + 2px))"

    #st.subheader("Progress")
    badge_html = []
    for i, title in enumerate(STEP_TITLES):
        if st.session_state.step > i:
            bg = BRAND_PRIME
            fg = WHITE
            border = BRAND_PRIME
        elif st.session_state.step == i:
            bg = WHITE
            fg = BLACK
            border = BRAND_PRIME
        else:
            bg = BRAND_GREY
            fg = WHITE
            border = BRAND_GREY

        left_pct = int(round((i / last_idx) * 100))
        if i == 0:
            transform = "translate(0, -50%)"
        elif i == len(STEP_TITLES) - 1:
            transform = "translate(-100%, -50%)"
        else:
            transform = "translate(-50%, -50%)"

        badge_html.append(
            textwrap.dedent(
                f"""
                <div style="
                    position:absolute;
                    top:50%;
                    left:{left_pct}%;
                    transform:{transform};
                    background:{bg};
                    color:{fg};
                    border:1px solid {border};
                    border-radius:999px;
                    padding:4px 10px;
                    font-size:0.82rem;
                    font-weight:600;
                    white-space:nowrap;
                    z-index:2;
                ">{i + 1}. {title}</div>
                """
            ).strip()
        )

    progress_html = textwrap.dedent(
        f"""
        <div style="position:relative; width:100%; height:52px; margin-top:4px;">
            <div style="
                position:absolute;
                left:0;
                right:0;
                top:50%;
                transform:translateY(-50%);
                height:14px;
                background:{BRAND_LIGHT};
                border-radius:999px;
                overflow:hidden;
                z-index:1;
            ">
                <div style="width:{fill_width_css}; height:100%; background:{BRAND_PRIME}; border-radius:999px;"></div>
            </div>
            {''.join(badge_html)}
        </div>
        """
    ).strip()

    st.markdown(
        progress_html,
        unsafe_allow_html=True,
    )


def render_back_next(can_next=True):
    c1, c_spacer, c2 = st.columns([1, 10, 1])
    with c1:
        st.button("Back", on_click=go_back, disabled=st.session_state.step == 0)
    with c2:
        next_inner_spacer, next_inner = st.columns([1, 1])
        with next_inner:
            st.button("Next", on_click=go_next, disabled=not can_next)


def run_processing_pipeline():
    tmp_dir = tempfile.mkdtemp(prefix="procard_app_")
    workflow_bytes = io.BytesIO(st.session_state.workflow_file.getvalue())
    output_file = os.path.join(tmp_dir, "Final_ProCard_Upload.xlsx")

    bank_zip = None
    bank_pdf_files = None
    if st.session_state.input_mode == "ZIP file":
        bank_zip = io.BytesIO(st.session_state.zip_file.getvalue())
    else:
        bank_pdf_files = [
            (uploaded.name, uploaded.getvalue())
            for uploaded in (st.session_state.pdf_files or [])
            if uploaded is not None
        ]

    log_buf = io.StringIO()
    with redirect_stdout(log_buf):
        summary = run_procard_pipeline(
            workflow_csv=workflow_bytes,
            bank_zip=bank_zip,
            bank_pdf_files=bank_pdf_files,
            batch_date=st.session_state.batch_date,
            manual_review_duplicates=False,
            prompt_for_missing_foapal=False,
            output_file=output_file,
        )

    output_file = summary.get("output_file", "Final_ProCard_Upload.xlsx")
    if not os.path.isabs(output_file):
        output_file = os.path.join(os.getcwd(), output_file)

    st.session_state.processing_log = log_buf.getvalue()
    st.session_state.summary = summary
    st.session_state.source_output_file = output_file
    st.session_state.teams_workflow = pd.read_excel(output_file, sheet_name="TEAMS WORKFLOW")
    st.session_state.teams_workflow_original = st.session_state.teams_workflow.copy(deep=True)
    st.session_state.bank_review = pd.read_excel(output_file, sheet_name="PDF Parsed Transactions")
    st.session_state.workflow_review = pd.read_excel(output_file, sheet_name="Original Workflow")
    st.session_state.workflow_only_rows = pd.read_excel(output_file, sheet_name="Workflow Not In PDF")
    st.session_state.foapal_review_df = None
    st.session_state.foapal_current_pos = 0
    st.session_state.foapal_review_skipped = False
    st.session_state.final_output_bytes = None
    st.session_state.final_output_name = ""
    st.session_state.final_output_signature = None
    st.session_state.processing_done = True


def build_final_workbook_bytes():
    teams_workflow = sanitize_for_output(st.session_state.teams_workflow)
    file_feed = build_file_feed_from_teams_workflow(teams_workflow, st.session_state.batch_date)
    file_feed = sanitize_for_output(file_feed)
    bank_review = sanitize_for_output(st.session_state.bank_review)
    workflow_review = sanitize_for_output(st.session_state.workflow_review)
    workflow_only_rows = sanitize_for_output(st.session_state.workflow_only_rows)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        file_feed.to_excel(writer, sheet_name="FILE FEED", index=False, header=False)
        teams_workflow.to_excel(writer, sheet_name="TEAMS WORKFLOW", index=False)
        bank_review.to_excel(writer, sheet_name="PDF Parsed Transactions", index=False)
        workflow_review.to_excel(writer, sheet_name="Original Workflow", index=False)
        workflow_only_rows.to_excel(writer, sheet_name="Workflow Not In PDF", index=False)

        file_feed_ws = writer.sheets.get("FILE FEED")
        if file_feed_ws is not None:
            for col_letter, width in zip(["A", "B", "C", "D", "E", "F", "G"], FILE_FEED_COLUMN_WIDTHS):
                file_feed_ws.column_dimensions[col_letter].width = width
        apply_workbook_font(writer.book)
    output.seek(0)
    return output.getvalue()


def load_from_previous_download(xlsx_bytes, batch_date):
    """Populate session state from a previously downloaded ProCard workbook and jump to FOAPAL review."""
    xls = pd.ExcelFile(io.BytesIO(xlsx_bytes))
    teams_workflow = pd.read_excel(xls, sheet_name="TEAMS WORKFLOW")
    bank_review = pd.read_excel(xls, sheet_name="PDF Parsed Transactions")
    workflow_review = pd.read_excel(xls, sheet_name="Original Workflow")
    workflow_only_rows = pd.read_excel(xls, sheet_name="Workflow Not In PDF")
    st.session_state.teams_workflow = teams_workflow
    st.session_state.teams_workflow_original = teams_workflow.copy(deep=True)
    st.session_state.bank_review = bank_review
    st.session_state.workflow_review = workflow_review
    st.session_state.workflow_only_rows = workflow_only_rows
    st.session_state.batch_date = batch_date
    st.session_state.processing_done = True
    st.session_state.processing_log = ""
    st.session_state.summary = {}
    st.session_state.source_output_file = ""
    st.session_state.foapal_review_df = None
    st.session_state.foapal_current_pos = 0
    st.session_state.foapal_review_skipped = False
    st.session_state.final_output_bytes = None
    st.session_state.final_output_name = ""
    st.session_state.final_output_signature = None
    st.session_state.step = 3


def build_final_output_signature():
    if st.session_state.teams_workflow is None:
        return None
    teams_for_sig = sanitize_for_output(st.session_state.teams_workflow)
    data_hash = int(pd.util.hash_pandas_object(teams_for_sig, index=True).sum())
    return (
        tuple(FILE_FEED_COLUMN_WIDTHS),
        st.session_state.batch_date,
        data_hash,
        bool(st.session_state.foapal_review_skipped),
    )


init_state()
apply_brand_theme()

header_logo_col, header_title_col = st.columns([1, 7])
with header_logo_col:
    render_top_logo()
with header_title_col:
    st.title("ProCard Reconciliation Web App")
render_step_nav()

step = st.session_state.step

if step == 0:
    st.subheader("Welcome")
    st.write(
        "Welcome to the ProCard reconciliation app. Upload a workflow CSV and a ZIP of statement PDFs, "
        "then walk through parsing, matching, FOAPAL review, and final download."
    )
    render_back_next(can_next=True)

elif step == 1:
    st.subheader("Upload Files")

    mode_box = st.container(key="mode_box")
    with mode_box:
        session_mode = st.radio(
            "Mode",
            ["New reconciliation", "Resume previous download"],
            key="session_mode",
            horizontal=True,
        )

    if session_mode == "Resume previous download":
        resume_upload_box = st.container(key="resume_upload_box")
        with resume_upload_box:
            resume_file = st.file_uploader(
                "Upload a previously downloaded ProCard workbook (.xlsx)",
                type=["xlsx"],
                key="resume_upload",
                help="Upload a Final_ProCard_Upload file previously downloaded from this app to continue reviewing unfilled FOAPAL rows.",
            )
        st.session_state.resume_file = resume_file

        st.markdown("<div style='height: 0.5rem;'></div>", unsafe_allow_html=True)
        resume_batch_date_box = st.container(key="resume_batch_date_box")
        with resume_batch_date_box:
            st.session_state.batch_date = st.text_input(
                "Batch date (last day of month, YYMMDD)",
                value=st.session_state.batch_date,
            )
            if st.session_state.batch_date and not is_last_day_batch_date_YYMMDD(st.session_state.batch_date):
                render_helper_box("Batch date should be the last day of the month in YYMMDD format.", "warning")

        can_resume = False
        if resume_file:
            try:
                xls = pd.ExcelFile(io.BytesIO(resume_file.getvalue()))
                required_sheets = {"TEAMS WORKFLOW", "PDF Parsed Transactions", "Original Workflow", "Workflow Not In PDF"}
                missing_sheets = required_sheets - set(xls.sheet_names)
                if missing_sheets:
                    render_helper_box(
                        f"Missing required sheets: {', '.join(sorted(missing_sheets))}. "
                        "Make sure this file was downloaded from the ProCard Reconciliation App.",
                        "error",
                    )
                else:
                    render_helper_box("Valid ProCard workbook detected. Click Next to continue FOAPAL review.", "success")
                    st.markdown("<div style='height: 0.35rem;'></div>", unsafe_allow_html=True)
                    can_resume = True
            except Exception:
                render_helper_box("Could not read this file. Make sure it is a valid .xlsx ProCard App download.", "error")

        c1, c_spacer, c2 = st.columns([1, 10, 1])
        with c1:
            st.button("Back", on_click=go_back, disabled=st.session_state.step == 0)
        with c2:
            next_inner_spacer, next_inner = st.columns([1, 1])
            with next_inner:
                st.button("Next", disabled=not can_resume, key="resume_next", on_click=request_resume_load)
        if can_resume and st.session_state.resume_requested and resume_file is not None:
            st.session_state.resume_requested = False
            load_from_previous_download(resume_file.getvalue(), st.session_state.batch_date)
            st.rerun()
        st.stop()

    # ── New reconciliation ────────────────────────────────────────────────────
    workflow_upload_box = st.container(key="workflow_upload_box")
    with workflow_upload_box:
        st.session_state.workflow_file = st.file_uploader(
            "Upload ProCard workflow CSV",
            type=["csv"],
            key="workflow_upload",
        )

    st.markdown("<div style='height: 0.5rem;'></div>", unsafe_allow_html=True)
    available_input_modes = ["ZIP file", "PDF files"]
    if st.session_state.input_mode not in available_input_modes:
        st.session_state.input_mode = "ZIP file"

    bank_upload_box = st.container(key="bank_upload_box")
    with bank_upload_box:
        input_mode = st.radio(
            "Bank statement upload method",
            available_input_modes,
            key="input_mode",
            horizontal=True,
        )
        if input_mode == "ZIP file":
            st.session_state.zip_file = st.file_uploader(
                "Upload ZIP containing bank statement PDFs",
                type=["zip"],
                key="zip_upload",
            )
            st.session_state.pdf_files = []
            bank_ready = st.session_state.zip_file is not None
        else:
            st.session_state.zip_file = None
            st.session_state.pdf_files = st.file_uploader(
                "Upload one or more bank statement PDFs",
                type=["pdf"],
                accept_multiple_files=True,
                key="pdf_uploads",
            )
            pdf_count = len(st.session_state.pdf_files or [])
            bank_ready = pdf_count > 0
            if pdf_count > 0:
                render_helper_box(f"{pdf_count} PDF file(s) uploaded.", "success")
                st.markdown("<div style='height: 0.35rem;'></div>", unsafe_allow_html=True)

    st.markdown("<div style='height: 0.5rem;'></div>", unsafe_allow_html=True)
    batch_date_box = st.container(key="batch_date_box")
    with batch_date_box:
        st.session_state.batch_date = st.text_input(
            "Batch date (last day of month, YYMMDD)",
            value=st.session_state.batch_date,
        )
        if st.session_state.batch_date and not is_last_day_batch_date_YYMMDD(st.session_state.batch_date):
            render_helper_box("Batch date should be the last day of the month in YYMMDD format.", "warning")

    can_next = st.session_state.workflow_file is not None and bank_ready
    render_back_next(can_next=can_next)

elif step == 2:
    st.subheader("Parse & Match")
    if not st.session_state.processing_done:
        parse_info_box = st.empty()
        render_helper_box("Click the button below to begin parsing and matching.", "info", target=parse_info_box)
        st.markdown("<div style='height: 0.45rem;'></div>", unsafe_allow_html=True)
        st.button(
            "Start Parsing and Matching",
            key="start_parse_btn",
            on_click=request_parse_run,
            disabled=st.session_state.parse_running,
        )
        if st.session_state.parse_requested and not st.session_state.parse_running:
            st.session_state.parse_requested = False
            st.session_state.parse_running = True
            parse_info_box.empty()
            try:
                with st.spinner("Processing files..."):
                    run_processing_pipeline()
            finally:
                st.session_state.parse_running = False
            st.rerun()

    if st.session_state.processing_done:
        summary = st.session_state.summary or {}

        bank_review = st.session_state.bank_review if st.session_state.bank_review is not None else pd.DataFrame()
        workflow_review = st.session_state.workflow_review if st.session_state.workflow_review is not None else pd.DataFrame()
        workflow_only = st.session_state.workflow_only_rows if st.session_state.workflow_only_rows is not None else pd.DataFrame()

        bank_tx_count = len(bank_review)
        workflow_matched_count = int(summary.get('workflow_matched_count', 0))
        workflow_previous_cycle_count = int(summary.get('workflow_previous_cycle_count', 0))

        pdf_total = float(summary.get('pdf_parsed_total', 0.0))
        comparable_workflow_total = float(summary.get('comparable_workflow_total', 0.0))
        previous_cycle_total = abs(float(summary.get('excluded_previous_statement_total', 0.0)))
        total_difference = float(summary.get('total_difference', pdf_total - comparable_workflow_total))
        rows_after_split = int(summary.get('rows_after_split', 0))
        statement_cycle = str(summary.get('statement_cycle', '') or '').strip()

        workflow_only_in_scope = pd.DataFrame(columns=workflow_only.columns)
        if not workflow_only.empty:
            warning_series = workflow_only['Warning Missing'].fillna('').astype(str) if 'Warning Missing' in workflow_only.columns else pd.Series(["" for _ in range(len(workflow_only))])
            out_of_date_mask = warning_series.str.startswith('Likely on previous statement') | warning_series.str.startswith('Likely on next statement')
            workflow_only_in_scope = workflow_only.loc[~out_of_date_mask].copy()

        totals_match = abs(total_difference) < 0.005
        if totals_match and workflow_only_in_scope.empty:
            render_helper_box("Processing complete. No errors detected.", "success")
        elif totals_match and not workflow_only_in_scope.empty:
            render_helper_box(
                "Processing complete, but one or more in-cycle Workflow rows are missing PDF matches. Review the rows below before proceeding.",
                "error",
            )
        else:
            render_helper_box("Processing complete. Errors detected: totals do not match.", "error")

        if statement_cycle:
            st.markdown("<div style='height: 0.35rem;'></div>", unsafe_allow_html=True)
            render_helper_box(
                f"<strong>Reconciliation Snapshot</strong><br>Current Billing Cycle (from bank statements): {statement_cycle}",
                "prime",
            )
            st.markdown("<div style='height: 0.75rem;'></div>", unsafe_allow_html=True)

        if not workflow_only_in_scope.empty:
            st.markdown("<div style='height: 0.5rem;'></div>", unsafe_allow_html=True)
            st.write("Workflow rows missing PDF match (in statement cycle)")
            st.dataframe(workflow_only_in_scope.fillna("").astype(str), width='stretch')

        if not totals_match:
            bank_cmp = bank_review.copy()
            workflow_cmp = workflow_review.copy()

            if not bank_cmp.empty:
                bank_cmp['_key'] = bank_cmp.apply(
                    lambda r: build_row_key_amount_date(r.get('Date', ''), r.get('Amount', '')),
                    axis=1,
                )
            if not workflow_cmp.empty:
                workflow_cmp['_key'] = workflow_cmp.apply(
                    lambda r: build_row_key_amount_date(r.get('Date', ''), r.get('Amount', '')),
                    axis=1,
                )

            if not workflow_only.empty and not workflow_cmp.empty:
                excluded_outdated_counter = Counter(
                    workflow_only.loc[
                        workflow_only['Warning Missing'].fillna('').astype(str).str.startswith('Likely on previous statement') |
                        workflow_only['Warning Missing'].fillna('').astype(str).str.startswith('Likely on next statement')
                    ].apply(lambda r: build_row_key_amount_date(r.get('Date', ''), r.get('Amount', '')), axis=1).tolist()
                )
                if excluded_outdated_counter:
                    drop_indices = []
                    for key, count in excluded_outdated_counter.items():
                        if count <= 0:
                            continue
                        key_indices = workflow_cmp.index[workflow_cmp['_key'] == key].tolist()
                        drop_indices.extend(key_indices[:int(count)])
                    if drop_indices:
                        workflow_cmp = workflow_cmp.drop(index=drop_indices).copy()

            bank_counter = Counter(bank_cmp['_key'].tolist()) if '_key' in bank_cmp.columns else Counter()
            workflow_counter = Counter(workflow_cmp['_key'].tolist()) if '_key' in workflow_cmp.columns else Counter()

            bank_only_counter = bank_counter - workflow_counter

            bank_only_rows = select_rows_by_key_counts(bank_cmp, '_key', bank_only_counter)
            if '_key' in bank_only_rows.columns:
                bank_only_rows = bank_only_rows.drop(columns=['_key'])

            if not bank_only_rows.empty:
                st.markdown("<div style='height: 0.5rem;'></div>", unsafe_allow_html=True)
                st.write("Mismatched PDF rows not found in Workflow")
                st.dataframe(bank_only_rows.fillna("").astype(str), width='stretch')

            if workflow_only_in_scope.empty and bank_only_rows.empty:
                render_helper_box("Totals differ, but no in-scope row-level mismatches were identified by key match.", "warning")

        m1, m2, m3 = st.columns(3)
        m1.metric("PDF Transactions", f"{bank_tx_count:,}")
        m2.metric(
            "Workflow Rows Matched to PDFs",
            f"{workflow_matched_count:,}",
            delta=f"-{workflow_previous_cycle_count:,} prior cycle" if workflow_previous_cycle_count else None,
            delta_color="off",
        )
        m3.metric("Rows After Split Expansion", f"{rows_after_split:,}")

        t1, t2, t3 = st.columns(3)
        t1.metric("PDF Total", f"${pdf_total:,.2f}")
        t2.metric("Workflow Total (Comparable)", f"${comparable_workflow_total:,.2f}")
        t3.metric("Amount Difference (PDF - Workflow)", f"${total_difference:,.2f}")

    render_back_next(can_next=st.session_state.processing_done)

elif step == 3:
    st.subheader("FOAPAL Review")
    if st.session_state.teams_workflow is None:
        render_helper_box("Run Parse & Match first.", "warning")
        st.markdown("<div style='height: 0.45rem;'></div>", unsafe_allow_html=True)
        render_back_next(can_next=False)
    else:
        required_cols = ['Fund Code:', 'Organization:', 'Account:', 'Program:', 'AD Code:']
        st.session_state.teams_workflow = normalize_editable_foapal(
            st.session_state.teams_workflow,
            required_cols,
        )
        foapal_option_source = (
            st.session_state.teams_workflow_original
            if st.session_state.teams_workflow_original is not None
            else st.session_state.teams_workflow
        )
        dropdown_options = build_foapal_dropdown_options(foapal_option_source, required_cols)
        missing_ids = list(st.session_state.teams_workflow.index[missing_required_mask(
            st.session_state.teams_workflow,
            required_cols,
        )])
        clamp_review_position(missing_ids)
        if not missing_ids:
            st.session_state.foapal_next_confirm = False

        if not missing_ids:
            render_helper_box("No missing FOAPAL values found.", "success")
            st.markdown("<div style='height: 0.45rem;'></div>", unsafe_allow_html=True)
        else:
            current_row_id = missing_ids[st.session_state.foapal_current_pos]
            current_row = st.session_state.teams_workflow.loc[current_row_id]

            render_helper_box(f"{len(missing_ids)} row(s) still need FOAPAL values.", "warning")
            st.markdown("<div style='height: 0.35rem;'></div>", unsafe_allow_html=True)
            st.caption(f"Reviewing row {st.session_state.foapal_current_pos + 1} of {len(missing_ids)}")

            st.markdown("<div style='height: 0.2rem;'></div>", unsafe_allow_html=True)

            nav_col1, nav_spacer, nav_col2 = st.columns([1.3, 8.4, 1.3])
            with nav_col1:
                if st.button(
                    "◀ Prev Row",
                    disabled=st.session_state.foapal_current_pos == 0,
                    key="foapal_prev_row",
                ):
                    st.session_state.foapal_next_confirm = False
                    st.session_state.foapal_current_pos -= 1
                    st.rerun()
            with nav_col2:
                next_row_spacer, next_row_inner = st.columns([1, 1])
                with next_row_inner:
                    if st.button(
                        "Next Row ▶",
                        disabled=st.session_state.foapal_current_pos >= len(missing_ids) - 1,
                        key="foapal_next_row",
                    ):
                        st.session_state.foapal_next_confirm = False
                        st.session_state.foapal_current_pos += 1
                        st.rerun()

            st.markdown("<div style='height: 0.2rem;'></div>", unsafe_allow_html=True)

            info_cols = st.columns(4)
            info_cols[0].metric("Date", str(current_row.get('Date of Transaction:', '')))
            info_cols[1].metric("Merchant", str(current_row.get('Merchant:', '')))
            info_cols[2].metric("Card Holder", str(current_row.get('Card Holder:', '')))
            info_cols[3].metric("Amount", str(current_row.get('Amount: $', '')))

            st.markdown("<div style='height: 0.2rem;'></div>", unsafe_allow_html=True)

            with st.form(f"foapal_form_{current_row_id}"):
                col1, col2, col3, col4, col5 = st.columns(5)
                with col1:
                    fund_code = selectbox_value(
                        "Fund",
                        current_row.get('Fund Code:', ''),
                        dropdown_options['Fund Code:'],
                    )
                with col2:
                    organization = selectbox_value(
                        "Org",
                        current_row.get('Organization:', ''),
                        dropdown_options['Organization:'],
                    )
                with col3:
                    account = selectbox_value(
                        "Account",
                        current_row.get('Account:', ''),
                        dropdown_options['Account:'],
                    )
                with col4:
                    ad_code = selectbox_value(
                        "AD Code",
                        current_row.get('AD Code:', ''),
                        dropdown_options['AD Code:'],
                    )
                with col5:
                    program = selectbox_value(
                        "Program",
                        current_row.get('Program:', ''),
                        dropdown_options['Program:'],
                        allow_outside_current=False,
                    )
                action_spacer, action_save = st.columns([5.2, 0.3])
                with action_save:
                    save_clicked = st.form_submit_button(
                        "Save This Row",
                        key="foapal_save_row_submit",
                        use_container_width=True,
                    )

            if save_clicked:
                updates = {
                    'Fund Code:': fund_code,
                    'Organization:': organization,
                    'Account:': account,
                    'Program:': program,
                    'AD Code:': ad_code,
                }
                for col, value in updates.items():
                    st.session_state.teams_workflow.at[current_row_id, col] = "" if is_blank(value) else str(value).strip()

                refreshed_missing_ids = list(st.session_state.teams_workflow.index[missing_required_mask(
                    st.session_state.teams_workflow,
                    required_cols,
                )])
                current_pos = st.session_state.foapal_current_pos
                row_still_missing = current_row_id in refreshed_missing_ids

                if not refreshed_missing_ids:
                    st.session_state.foapal_current_pos = 0
                elif not row_still_missing:
                    st.session_state.foapal_current_pos = min(current_pos, len(refreshed_missing_ids) - 1)

                st.session_state.foapal_next_confirm = False
                render_helper_box("FOAPAL saved.", "success")
                st.rerun()

        nav_back, nav_spacer, nav_next = st.columns([1, 10, 1])
        with nav_back:
            if st.button("Back", key="foapal_back"):
                st.session_state.foapal_next_confirm = False
                go_back()
                st.rerun()
        with nav_next:
            next_inner_spacer, next_inner = st.columns([1, 1])
            with next_inner:
                if st.button("Next", key="foapal_next_step"):
                    if missing_ids:
                        st.session_state.foapal_next_confirm = True
                        st.rerun()
                    go_next()
                    st.rerun()

        if missing_ids and st.session_state.foapal_next_confirm:
            render_foapal_missing_confirm_dialog(len(missing_ids))

elif step == 4:
    st.subheader("Download Output")
    if st.session_state.teams_workflow is None:
        render_helper_box("No processed data available yet.", "warning")
        st.markdown("<div style='height: 0.45rem;'></div>", unsafe_allow_html=True)
        render_back_next(can_next=False)
    else:
        required_cols = ['Fund Code:', 'Organization:', 'Account:', 'Program:', 'AD Code:']
        has_missing_required = missing_required_mask(st.session_state.teams_workflow, required_cols).any()
        if has_missing_required:
            render_helper_box("Some FOAPAL values are missing. You can still download the output and fill them in later.", "warning")
            st.markdown("<div style='height: 0.45rem;'></div>", unsafe_allow_html=True)

        current_output_signature = build_final_output_signature()
        if (
            not st.session_state.final_output_bytes
            or st.session_state.final_output_signature != current_output_signature
        ):
            st.session_state.final_output_bytes = build_final_workbook_bytes()
            st.session_state.final_output_name = f"Final_ProCard_Upload_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            st.session_state.final_output_signature = current_output_signature

        st.download_button(
            label="Download Final Workbook",
            data=st.session_state.final_output_bytes,
            file_name=st.session_state.final_output_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            on_click="ignore",
        )

        st.write("FILE FEED Preview")
        file_feed_preview = build_file_feed_from_teams_workflow(
            sanitize_for_output(st.session_state.teams_workflow),
            st.session_state.batch_date,
        )
        st.dataframe(sanitize_for_output(file_feed_preview), width='stretch')
        render_back_next(can_next=False)